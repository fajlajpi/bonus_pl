"""
Template Text Extraction Script for Django Projects
====================================================
Walks through Django template directories and extracts all human-readable text
from HTML templates, outputting an Excel workbook suitable for translation workflows.

Usage:
    python extract_template_strings.py <template_dir> [--output filename.xlsx] [--exclude-dirs dir1,dir2]

Example:
    python extract_template_strings.py pa_bonus/templates --output translations.xlsx --exclude-dirs manager

The script extracts:
    - Plain text content from HTML elements (headings, paragraphs, links, labels, etc.)
    - Text from Django template tags like {% block title %}
    - Button and input labels
    - Alt text and placeholder attributes
    - Option text in select elements

The script ignores:
    - Django template variables ({{ variable }})
    - Template logic tags ({% if %}, {% for %}, {% csrf_token %}, etc.)
    - HTML comments
    - Pure whitespace or numeric-only strings
    - JavaScript and CSS blocks
    - Manager/admin templates (configurable via --exclude-dirs)

Output columns:
    A: Template File     - relative path to the template
    B: Line (approx.)    - approximate line number in the source file
    C: Context           - the HTML tag or attribute where the text appears
    D: Original Text     - the extracted Czech/source text
    E: Machine Translation - empty column for machine translation
    F: Human Translation   - empty column for human review
    G: Notes             - empty column for translator notes
"""

import argparse
import os
import re
from pathlib import Path

from bs4 import BeautifulSoup, Comment, NavigableString
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# --- Configuration ---

# Tags whose text content we want to extract
CONTENT_TAGS = {
    'h1', 'h2', 'h3', 'h4', 'h5', 'h6',
    'p', 'a', 'span', 'strong', 'em', 'b', 'i',
    'li', 'td', 'th', 'label', 'button',
    'small', 'legend', 'figcaption', 'dt', 'dd',
    'option', 'summary',
}

# Attributes that may contain translatable text
TRANSLATABLE_ATTRS = {
    'alt', 'title', 'placeholder', 'aria-label', 'value',
}

# Tags to skip entirely (their children too)
SKIP_TAGS = {'script', 'style', 'code', 'pre', 'svg'}

# Django template tags that are purely structural / logic (not text)
LOGIC_TAG_PATTERN = re.compile(
    r'\{%\s*(?:if|elif|else|endif|for|endfor|empty|endblock|block|extends|load|'
    r'include|csrf_token|url|static|with|endwith|comment|endcomment|'
    r'firstof|cycle|regroup|spaceless|endspaceless|verbatim|endverbatim|'
    r'autoescape|endautoescape|filter|endfilter|now|widthratio|'
    r'templatetag|debug|resetcycle)\b[^%]*%\}'
)

# Pattern for Django variables {{ ... }}
VARIABLE_PATTERN = re.compile(r'\{\{[^}]+\}\}')

# Pattern for Django template tags {% ... %}
TEMPLATE_TAG_PATTERN = re.compile(r'\{%[^%]+%\}')

# Pattern for block title content: {% block title %} Some Text {% endblock %}
BLOCK_TITLE_PATTERN = re.compile(
    r'\{%\s*block\s+title\s*%\}\s*(.+?)\s*\{%\s*endblock',
    re.DOTALL
)

# Pattern for {% trans "..." %} tags (already i18n-wrapped, still useful to capture)
TRANS_TAG_PATTERN = re.compile(r'\{%\s*trans\s+["\'](.+?)["\']\s*%\}')


def is_translatable(text: str) -> bool:
    """
    Determine whether a string is worth translating.
    Filters out empty strings, pure numbers, lone punctuation, 
    template variables standing alone, and very short non-word content.
    """
    if not text or not text.strip():
        return False

    cleaned = text.strip()

    # Skip if it is purely a Django variable or tag
    if re.fullmatch(r'\{\{[^}]+\}\}', cleaned):
        return False
    if re.fullmatch(r'\{%[^%]+%\}', cleaned):
        return False

    # Strip out Django variables and tags, then check what remains
    remainder = VARIABLE_PATTERN.sub('', cleaned)
    remainder = TEMPLATE_TAG_PATTERN.sub('', remainder)
    remainder = remainder.strip()

    # If nothing meaningful remains after removing template syntax, skip
    if not remainder:
        return False

    # Skip pure numbers, dates, or punctuation
    if re.fullmatch(r'[\d\s\.,:/\-\+%]+', remainder):
        return False

    # Skip very short strings that are just symbols
    if len(remainder) <= 1 and not remainder.isalpha():
        return False

    return True


def clean_text(text: str) -> str:
    """
    Clean extracted text for presentation in the spreadsheet.
    Normalizes whitespace while preserving Django template syntax for context.
    """
    # Collapse internal whitespace
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def extract_block_titles(raw_html: str, filepath: str) -> list[dict]:
    """
    Extract text from {% block title %} ... {% endblock %} patterns.
    These are not visible in the parsed DOM because BeautifulSoup does not
    understand Django template tags natively.
    """
    results = []
    for match in BLOCK_TITLE_PATTERN.finditer(raw_html):
        text = match.group(1).strip()
        # Remove any nested template tags
        text = TEMPLATE_TAG_PATTERN.sub('', text).strip()
        if is_translatable(text):
            # Estimate line number
            line_no = raw_html[:match.start()].count('\n') + 1
            results.append({
                'file': filepath,
                'line': line_no,
                'context': '{% block title %}',
                'text': clean_text(text),
            })
    return results


def extract_trans_tags(raw_html: str, filepath: str) -> list[dict]:
    """
    Extract text from {% trans "..." %} tags.
    These are already marked for translation but we still want them inventoried.
    """
    results = []
    for match in TRANS_TAG_PATTERN.finditer(raw_html):
        text = match.group(1).strip()
        if is_translatable(text):
            line_no = raw_html[:match.start()].count('\n') + 1
            results.append({
                'file': filepath,
                'line': line_no,
                'context': '{% trans %}',
                'text': clean_text(text),
            })
    return results


def extract_from_template(filepath: str, base_dir: str) -> list[dict]:
    """
    Parse a single Django template file and extract all translatable strings.
    
    The approach:
    1. Read the raw file and extract block titles and trans tags via regex.
    2. Strip Django template tags so BeautifulSoup can parse the HTML structure.
    3. Walk the DOM tree and collect text from content elements and attributes.
    """
    rel_path = os.path.relpath(filepath, base_dir)

    with open(filepath, 'r', encoding='utf-8') as f:
        raw_html = f.read()

    results = []

    # Phase 1: Regex-based extraction for Django-specific patterns
    results.extend(extract_block_titles(raw_html, rel_path))
    results.extend(extract_trans_tags(raw_html, rel_path))

    # Phase 2: Prepare HTML for BeautifulSoup by neutralizing Django tags
    # Replace template tags with empty strings for parsing purposes,
    # but keep the original for line number estimation.
    # We keep {{ variables }} as placeholders so we can see them in context.
    parseable_html = LOGIC_TAG_PATTERN.sub('', raw_html)
    # Replace remaining template tags (like {% url ... %}) with empty string
    parseable_html = re.sub(r'\{%[^%]*%\}', '', parseable_html)

    soup = BeautifulSoup(parseable_html, 'lxml')

    # Remove elements we do not care about
    for tag in soup.find_all(SKIP_TAGS):
        tag.decompose()

    # Remove HTML comments
    for comment in soup.find_all(string=lambda s: isinstance(s, Comment)):
        comment.extract()

    # Collect text from elements
    seen_texts = set()  # Deduplicate within a single file

    for tag in soup.find_all(CONTENT_TAGS):
        # Get direct text content (not from children tags)
        # We use .strings to get all text nodes, then join
        direct_texts = []
        for child in tag.children:
            if isinstance(child, NavigableString) and not isinstance(child, Comment):
                t = str(child).strip()
                if t:
                    direct_texts.append(t)

        if direct_texts:
            full_text = ' '.join(direct_texts)
            full_text = clean_text(full_text)

            if is_translatable(full_text) and full_text not in seen_texts:
                seen_texts.add(full_text)
                # Try to find approximate line number by searching raw HTML
                # This is a best-effort approach
                search_fragment = full_text[:40]  # Use first 40 chars
                line_no = 0
                idx = raw_html.find(search_fragment)
                if idx >= 0:
                    line_no = raw_html[:idx].count('\n') + 1

                results.append({
                    'file': rel_path,
                    'line': line_no if line_no > 0 else '?',
                    'context': f'<{tag.name}>',
                    'text': full_text,
                })

        # Check translatable attributes
        for attr in TRANSLATABLE_ATTRS:
            attr_val = tag.get(attr)
            if attr_val and isinstance(attr_val, str):
                attr_val = clean_text(attr_val)
                if is_translatable(attr_val) and attr_val not in seen_texts:
                    seen_texts.add(attr_val)
                    results.append({
                        'file': rel_path,
                        'line': '?',
                        'context': f'<{tag.name} {attr}="">',
                        'text': attr_val,
                    })

    return results


def build_workbook(all_results: list[dict], output_path: str) -> None:
    """
    Create a professionally formatted Excel workbook from extraction results.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Template Strings'

    # -- Styles --
    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2F5496')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(vertical='top', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    translation_fill = PatternFill('solid', fgColor='FFF2CC')  # Light yellow
    notes_fill = PatternFill('solid', fgColor='E2EFDA')        # Light green

    # -- Headers --
    headers = [
        ('Template File', 30),
        ('Line', 8),
        ('Context', 18),
        ('Original Text', 55),
        ('Machine Translation', 55),
        ('Human Translation', 55),
        ('Notes', 30),
    ]

    for col_idx, (title, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # -- Data rows --
    for row_idx, entry in enumerate(all_results, start=2):
        values = [
            entry['file'],
            entry['line'],
            entry['context'],
            entry['text'],
            '',  # Machine Translation
            '',  # Human Translation
            '',  # Notes
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

            # Highlight translation columns
            if col_idx == 5:
                cell.fill = translation_fill
            elif col_idx == 6:
                cell.fill = translation_fill
            elif col_idx == 7:
                cell.fill = notes_fill

    # Freeze top row and add auto-filter
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    # -- Summary sheet --
    summary = wb.create_sheet('Summary')
    summary_headers = [
        ('Metric', 30),
        ('Value', 20),
    ]
    for col_idx, (title, width) in enumerate(summary_headers, start=1):
        cell = summary.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        summary.column_dimensions[get_column_letter(col_idx)].width = width

    # Count unique files
    unique_files = set(e['file'] for e in all_results)
    summary_data = [
        ('Total translatable strings', len(all_results)),
        ('Template files processed', len(unique_files)),
        ('Files with translatable content', len(unique_files)),
    ]

    # Per-file breakdown
    file_counts = {}
    for entry in all_results:
        file_counts[entry['file']] = file_counts.get(entry['file'], 0) + 1

    for row_idx, (label, value) in enumerate(summary_data, start=2):
        summary.cell(row=row_idx, column=1, value=label).font = Font(name='Arial', size=10)
        summary.cell(row=row_idx, column=2, value=value).font = Font(name='Arial', size=10, bold=True)

    # File breakdown section
    breakdown_start = len(summary_data) + 3
    summary.cell(row=breakdown_start, column=1, value='Strings per File').font = Font(
        name='Arial', size=10, bold=True
    )
    for i, (filename, count) in enumerate(sorted(file_counts.items()), start=breakdown_start + 1):
        summary.cell(row=i, column=1, value=filename).font = Font(name='Arial', size=10)
        summary.cell(row=i, column=2, value=count).font = Font(name='Arial', size=10)

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description='Extract translatable text from Django templates into an Excel workbook.'
    )
    parser.add_argument(
        'template_dir',
        help='Path to the templates directory (e.g., pa_bonus/templates)'
    )
    parser.add_argument(
        '--output', '-o',
        default='template_strings_for_translation.xlsx',
        help='Output Excel file path (default: template_strings_for_translation.xlsx)'
    )
    parser.add_argument(
        '--exclude-dirs',
        default='',
        help='Comma-separated directory names to exclude (e.g., manager,admin)'
    )

    args = parser.parse_args()
    template_dir = args.template_dir
    output_path = args.output
    exclude_dirs = set(d.strip() for d in args.exclude_dirs.split(',') if d.strip())

    if not os.path.isdir(template_dir):
        print(f"Error: '{template_dir}' is not a valid directory.")
        return

    print(f"Scanning templates in: {template_dir}")
    if exclude_dirs:
        print(f"Excluding directories: {', '.join(exclude_dirs)}")

    all_results = []
    files_processed = 0

    for root, dirs, files in os.walk(template_dir):
        # Filter out excluded directories
        dirs[:] = [d for d in dirs if d not in exclude_dirs]

        for filename in sorted(files):
            if not filename.endswith('.html'):
                continue

            filepath = os.path.join(root, filename)
            entries = extract_from_template(filepath, template_dir)
            all_results.extend(entries)
            files_processed += 1

            if entries:
                print(f"  {os.path.relpath(filepath, template_dir)}: {len(entries)} strings")
            else:
                print(f"  {os.path.relpath(filepath, template_dir)}: (no translatable strings)")

    print(f"\nProcessed {files_processed} template files.")
    print(f"Found {len(all_results)} translatable strings.")

    if all_results:
        build_workbook(all_results, output_path)
        print(f"Output saved to: {output_path}")
    else:
        print("No translatable strings found. No output file created.")


if __name__ == '__main__':
    main()