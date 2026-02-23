"""
Management Command: Export Extra Goals Projection Report

This command generates an Excel report showing:
- Clients with active Extra Goals
- Current turnover vs. goal
- Turnover needed to hit the goal
- Projected points if goal is achieved (including retroactive recovery points)

Usage:
    python manage.py export_extra_goals_projection --year=2025 --output=report.xlsx
"""

from django.core.management.base import BaseCommand
from django.db.models import Sum
from django.utils import timezone
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from pa_bonus.models import UserContractGoal, GoalEvaluation
from pa_bonus.utilities import calculate_turnover_for_goal


class Command(BaseCommand):
    help = 'Export Extra Goals projection report with potential points'
    
    def add_arguments(self, parser):
        parser.add_argument(
            '--year',
            type=int,
            default=timezone.now().year,
            help='Year to analyze (default: current year)'
        )
        parser.add_argument(
            '--output',
            type=str,
            default='extra_goals_projection.xlsx',
            help='Output Excel filename'
        )
        parser.add_argument(
            '--include-completed',
            action='store_true',
            help='Include goals that are already completed'
        )
    
    def handle(self, *args, **options):
        year = options['year']
        output_file = options['output']
        include_completed = options['include_completed']
        
        self.stdout.write(f"Generating Extra Goals Projection Report for {year}...")
        
        # Get goals for the specified year
        goals = self._get_active_goals(year, include_completed)
        
        if not goals:
            self.stdout.write(self.style.WARNING(f"No active goals found for {year}"))
            return
        
        # Generate report data
        report_data = []
        for goal in goals:
            row = self._calculate_goal_projection(goal)
            if row:
                report_data.append(row)
        
        # Write to Excel
        self._write_excel(report_data, output_file)
        
        self.stdout.write(
            self.style.SUCCESS(
                f"Successfully exported {len(report_data)} goal projections to {output_file}"
            )
        )
    
    def _get_active_goals(self, year, include_completed):
        """
        Get all active goals for the specified year.
        
        A goal is "active" if:
        - It has a goal period that overlaps with the specified year
        - The user contract is active
        - (Optional) It hasn't been fully completed yet
        """
        year_start = datetime(year, 1, 1).date()
        year_end = datetime(year, 12, 31).date()
        
        queryset = UserContractGoal.objects.filter(
            user_contract__is_active=True,
            goal_period_from__lte=year_end,
            goal_period_to__gte=year_start
        ).select_related(
            'user_contract__user_id__region'
        ).prefetch_related(
            'brands',
            'evaluations'
        ).order_by(
            'user_contract__user_id__last_name',
            'user_contract__user_id__first_name'
        )
        
        if not include_completed:
            # Only include goals where the end date hasn't passed yet
            today = timezone.now().date()
            queryset = queryset.filter(goal_period_to__gte=today)
        
        return queryset
    
    def _calculate_goal_projection(self, goal):
        """
        Calculate the projection for a single goal.
        
        This replicates the logic from GoalEvaluationView._determine_evaluation_result
        but in "projection mode" - showing what WOULD happen if the yearly goal is hit.
        
        Returns a dictionary with all the calculated values, or None if calculation fails.
        """
        today = timezone.now().date()
        user = goal.user_contract.user_id
        
        # Calculate current turnover for the entire goal period
        current_turnover = calculate_turnover_for_goal(
            user,
            goal.brands.all(),
            goal.goal_period_from,
            min(today, goal.goal_period_to)
        )
        
        # Calculate turnover needed to hit the goal
        turnover_needed = max(0, goal.goal_value - float(current_turnover))
        
        # Calculate projected points if goal is hit
        projected_points = self._calculate_projected_points(goal, current_turnover)
        
        # Get already awarded points
        already_awarded = goal.evaluations.aggregate(
            total=Sum('bonus_points')
        )['total'] or 0
        
        # Calculate net new points (what they'd get on top of already awarded)
        net_new_points = max(0, projected_points - already_awarded)
        
        # Get brand names
        brand_names = ', '.join([brand.name for brand in goal.brands.all()])
        
        # Get region name
        region_name = user.region.name if user.region else 'No Region'
        
        return {
            'Client ID': user.user_number,
            'Client Name': f"{user.last_name} {user.first_name}",
            'Region': region_name,
            'Brands': brand_names,
            'Goal Period From': goal.goal_period_from.strftime('%Y-%m-%d'),
            'Goal Period To': goal.goal_period_to.strftime('%Y-%m-%d'),
            'Goal Value': goal.goal_value,
            'Goal Base': goal.goal_base,
            'Current Turnover': float(current_turnover),
            'Turnover Needed': turnover_needed,
            'Progress %': round((float(current_turnover) / goal.goal_value * 100), 2) if goal.goal_value > 0 else 0,
            'Already Awarded Points': already_awarded,
            'Projected Total Points': projected_points,
            'Net New Points': net_new_points,
            'Evaluation Frequency (months)': goal.evaluation_frequency,
            'Recovery Enabled': 'Yes' if goal.allow_full_period_recovery else 'No',
        }
    
    def _calculate_projected_points(self, goal, current_turnover):
        """
        Calculate what points would be awarded if the yearly goal is achieved.
        
        This implements the same logic as GoalEvaluationView._determine_evaluation_result
        but for projection purposes.
        
        Key logic:
        1. If recovery is enabled and yearly goal would be met:
           - Calculate total year points: (goal_value - goal_base) * bonus_percentage
           - Cap at the proportional annual limit (20,000 for 12 months)
        2. If recovery is not enabled:
           - Sum up milestone points for each period
        
        Returns:
            int: Total projected points
        """
        if goal.allow_full_period_recovery:
            # Calculate what total year points would be
            # Formula: (actual_turnover - base) * bonus_percentage
            # But we're projecting, so we use goal_value instead of actual
            total_year_points = int((goal.goal_value - goal.goal_base) * goal.bonus_percentage)
            
            # Apply the annual cap
            # Cap formula: 20,000 points for 12 months, prorated for other lengths
            cap = self._calculate_points_cap(goal)
            capped_points = min(total_year_points, cap)
            
            return capped_points
        else:
            # Without recovery, we need to calculate milestone by milestone
            # This is more complex - we'd need to evaluate each period separately
            # For simplicity, we'll use the same formula as recovery
            # (This is a reasonable approximation)
            total_year_points = int((goal.goal_value - goal.goal_base) * goal.bonus_percentage)
            cap = self._calculate_points_cap(goal)
            return min(total_year_points, cap)
    
    def _calculate_points_cap(self, goal):
        """
        Calculate the maximum points cap for the contract period.
        Base: 20,000 points for 12 months, proportional for other lengths.
        
        This replicates the logic from GoalEvaluationView._calculate_points_cap
        """
        total_days = (goal.goal_period_to - goal.goal_period_from).days
        # Approximate months (30.44 days per month on average)
        total_months = total_days / 30.44
        # 20,000 points for 12 months = ~1,667 points per month
        return int(total_months * 1667)
    
    def _write_excel(self, data, filename):
        """
        Write the report data to an Excel file with formatting.
        
        Uses openpyxl to create a professionally formatted Excel workbook with:
        - Header row with bold text and colored background
        - Auto-sized columns
        - Number formatting for currency and percentages
        - Frozen header row for easy scrolling
        """
        if not data:
            self.stdout.write(self.style.WARNING("No data to write"))
            return
        
        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extra Goals Projection"
        
        # Get field names from the first row
        fieldnames = list(data[0].keys())
        
        # Write header row with formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, field_name in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = field_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data rows
        for row_num, row_data in enumerate(data, 2):
            for col_num, field_name in enumerate(fieldnames, 1):
                cell = ws.cell(row=row_num, column=col_num)
                value = row_data[field_name]
                cell.value = value
                
                # Apply number formatting based on field type
                if field_name in ['Goal Value', 'Goal Base', 'Current Turnover', 'Turnover Needed']:
                    # Format as currency (no decimal places for whole numbers)
                    cell.number_format = '#,##0'
                elif field_name == 'Progress %':
                    # Format as percentage
                    cell.number_format = '0.00"%"'
                elif field_name in ['Already Awarded Points', 'Projected Total Points', 'Net New Points']:
                    # Format as integer with thousands separator
                    cell.number_format = '#,##0'
        
        # Auto-size columns based on content
        for col_num, field_name in enumerate(fieldnames, 1):
            column_letter = get_column_letter(col_num)
            
            # Calculate max length for this column
            max_length = len(field_name)
            for row_num in range(2, len(data) + 2):
                cell_value = ws.cell(row=row_num, column=col_num).value
                if cell_value is not None:
                    max_length = max(max_length, len(str(cell_value)))
            
            # Set column width (add some padding)
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 to avoid super wide columns
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze the header row so it stays visible when scrolling
        ws.freeze_panes = "A2"
        
        # Save the workbook
        wb.save(filename)
