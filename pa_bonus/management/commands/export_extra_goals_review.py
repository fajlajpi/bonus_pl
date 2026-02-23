"""
Management Command: Export Extra Goals Year-End Review

Generates a retrospective Excel report for all clients who had Extra Goals
set during a given calendar year (defaults to the previous year).

Columns produced:
    - Client number (ZC), name (Last, First), region, brands
    - Goal period start/end, base, goal value
    - Net turnover (invoices minus credit notes), total and per-brand
    - Credit note amounts (raw), total and per-brand
    - Standard invoice points (total)
    - Credit note point deductions (total)
    - Extra goal points awarded (from GoalEvaluation)

Usage:
    python manage.py export_extra_goals_review
    python manage.py export_extra_goals_review --year=2024
    python manage.py export_extra_goals_review --output=review_2025.xlsx
    python manage.py export_extra_goals_review --include-pending
"""

from datetime import date
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from django.core.management.base import BaseCommand
from django.db.models import Sum, Value, DecimalField, Q
from django.db.models.functions import Coalesce
from django.utils import timezone

from pa_bonus.models import (
    UserContractGoal,
    GoalEvaluation,
    InvoiceBrandTurnover,
    PointsTransaction,
    Brand,
)

# ---------------------------------------------------------------------------
# Hard-coded brand names.  If more brands appear in the future you can
# simply extend this list.  The command looks up actual Brand objects by
# name so the spelling must match what is in the database exactly.
# ---------------------------------------------------------------------------
BRAND_NAMES = ["Echosline", "Alter Ego"]


class Command(BaseCommand):
    help = "Export a year-end review of Extra Goals to an Excel file"

    # ------------------------------------------------------------------
    # CLI arguments
    # ------------------------------------------------------------------
    def add_arguments(self, parser):
        parser.add_argument(
            "--year",
            type=int,
            default=timezone.now().year - 1,
            help="Calendar year to review (default: previous year)",
        )
        parser.add_argument(
            "--output",
            type=str,
            default="",
            help="Output .xlsx filename (default: extra_goals_review_<year>.xlsx)",
        )
        parser.add_argument(
            "--include-pending",
            action="store_true",
            help="Include PENDING points in addition to CONFIRMED",
        )

    # ------------------------------------------------------------------
    # Main entry point
    # ------------------------------------------------------------------
    def handle(self, *args, **options):
        year = options["year"]
        output_file = options["output"] or f"extra_goals_review_{year}.xlsx"
        self.include_pending = options["include_pending"]

        if self.include_pending:
            self.stdout.write("Including PENDING points in the report.")

        self.stdout.write(f"Generating Extra Goals Year-End Review for {year} ...")

        # Resolve Brand objects once so we can reuse them in queries.
        brand_map = {}
        for name in BRAND_NAMES:
            try:
                brand_map[name] = Brand.objects.get(name=name)
            except Brand.DoesNotExist:
                self.stderr.write(
                    self.style.WARNING(
                        f"Brand '{name}' not found in the database -- "
                        f"the corresponding columns will be empty."
                    )
                )
                brand_map[name] = None

        # ----- Fetch all goals that overlapped with the requested year -----
        # A goal "belongs" to the year if its period intersects [Jan 1 .. Dec 31].
        year_start = date(year, 1, 1)
        year_end = date(year, 12, 31)

        goals = (
            UserContractGoal.objects.filter(
                goal_period_from__lte=year_end,
                goal_period_to__gte=year_start,
            )
            .select_related("user_contract__user_id__region")
            .prefetch_related("brands", "evaluations")
            .order_by("user_contract__user_id__last_name")
        )

        if not goals.exists():
            self.stdout.write(
                self.style.WARNING(f"No Extra Goals found for {year}.")
            )
            return

        # ----- Build one row per goal -----
        rows = []
        for goal in goals:
            user = goal.user_contract.user_id
            row = self._build_row(goal, user, year_start, year_end, brand_map)
            rows.append(row)

        # ----- Write Excel -----
        self._write_excel(rows, output_file, year, brand_map)

        self.stdout.write(
            self.style.SUCCESS(
                f"Done -- {len(rows)} rows written to {output_file}"
            )
        )

    # ------------------------------------------------------------------
    # Build a single data row for one goal
    # ------------------------------------------------------------------
    def _build_row(self, goal, user, year_start, year_end, brand_map):
        """
        Collect every piece of data requested for a single goal.

        The date range we use for turnover / credit note / points queries is
        the intersection of the goal period and the calendar year, so we never
        count data that falls outside either boundary.
        """
        # Effective date range = intersection of goal period and calendar year
        eff_start = max(goal.goal_period_from, year_start)
        eff_end = min(goal.goal_period_to, year_end)

        brand_names = ", ".join(b.name for b in goal.brands.all())

        # -- Turnover (invoices) per brand and total ----------------------
        goal_brands = goal.brands.all()
        invoice_total, invoice_by_brand = self._turnover(
            user, eff_start, eff_end, brand_map, goal_brands, invoice_type="INVOICE"
        )

        # -- Credit notes per brand and total -----------------------------
        credit_total, credit_by_brand = self._turnover(
            user, eff_start, eff_end, brand_map, goal_brands, invoice_type="CREDIT_NOTE"
        )

        # -- Net turnover = invoices minus credit notes -------------------
        net_turnover_total = float(invoice_total) - float(credit_total)
        net_turnover_by_brand = {
            name: float(invoice_by_brand.get(name, 0)) - float(credit_by_brand.get(name, 0))
            for name in BRAND_NAMES
        }

        # -- Standard invoice points (STANDARD_POINTS) --------------------
        standard_points = self._sum_points(
            user, eff_start, eff_end, point_type="STANDARD_POINTS"
        )

        # -- Credit note point deductions (CREDIT_NOTE_ADJUST) ------------
        credit_note_points = self._sum_points(
            user, eff_start, eff_end, point_type="CREDIT_NOTE_ADJUST"
        )

        # -- Extra goal points from evaluations ---------------------------
        extra_points = self._extra_points_from_evaluations(goal)

        return {
            "client_number": user.user_number,
            "client_name": f"{user.last_name}, {user.first_name}",
            "region": user.region.name if user.region else "",
            "brands": brand_names,
            "goal_from": goal.goal_period_from,
            "goal_to": goal.goal_period_to,
            "goal_base": goal.goal_base,
            "goal_value": goal.goal_value,
            "turnover_total": net_turnover_total,
            **{
                f"turnover_{name}": net_turnover_by_brand[name]
                for name in BRAND_NAMES
            },
            "credit_total": float(credit_total),
            **{
                f"credit_{name}": float(credit_by_brand.get(name, 0))
                for name in BRAND_NAMES
            },
            "standard_points": standard_points,
            "credit_note_points": credit_note_points,
            "extra_points": extra_points,
        }

    # ------------------------------------------------------------------
    # Helper: turnover (or credit notes) per brand
    # ------------------------------------------------------------------
    def _turnover(self, user, start, end, brand_map, goal_brands, invoice_type):
        """
        Sum InvoiceBrandTurnover amounts for the given user, date range, and
        invoice type.  The total is restricted to the brands assigned to the
        goal (goal_brands), while the per-brand breakdown uses the hard-coded
        BRAND_NAMES for the dedicated columns.

        Returns (total, {brand_name: amount}).
        """
        base_qs = InvoiceBrandTurnover.objects.filter(
            invoice__client_number=user.user_number,
            invoice__invoice_date__gte=start,
            invoice__invoice_date__lte=end,
            invoice__invoice_type=invoice_type,
        )

        # Total: only brands that belong to this goal
        total = base_qs.filter(brand__in=goal_brands).aggregate(
            t=Coalesce(Sum("amount"), Value(0, output_field=DecimalField()))
        )["t"]

        by_brand = {}
        for name, brand_obj in brand_map.items():
            if brand_obj is None:
                by_brand[name] = Decimal("0")
                continue
            by_brand[name] = base_qs.filter(brand=brand_obj).aggregate(
                t=Coalesce(Sum("amount"), Value(0, output_field=DecimalField()))
            )["t"]

        return total, by_brand

    # ------------------------------------------------------------------
    # Helper: sum confirmed points of a given type
    # ------------------------------------------------------------------
    def _sum_points(self, user, start, end, point_type):
        """
        Sum PointsTransaction.value for the user in the date range.

        By default only CONFIRMED transactions are counted.  When the
        --include-pending flag is active, PENDING transactions are included
        as well.
        """
        allowed_statuses = ["CONFIRMED"]
        if self.include_pending:
            allowed_statuses.append("PENDING")

        return (
            PointsTransaction.objects.filter(
                user=user,
                date__gte=start,
                date__lte=end,
                type=point_type,
                status__in=allowed_statuses,
            ).aggregate(t=Coalesce(Sum("value"), Value(0)))["t"]
        )

    # ------------------------------------------------------------------
    # Helper: extra points from GoalEvaluation records
    # ------------------------------------------------------------------
    def _extra_points_from_evaluations(self, goal):
        """
        Total bonus_points recorded in GoalEvaluation rows linked to this
        goal.  These are the milestone / recovery / final extra points that
        were actually awarded.
        """
        return goal.evaluations.aggregate(
            t=Coalesce(Sum("bonus_points"), Value(0))
        )["t"]

    # ------------------------------------------------------------------
    # Excel generation
    # ------------------------------------------------------------------
    def _write_excel(self, rows, filename, year, brand_map):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Extra Goals {year}"

        # ---- Styles -----------------------------------------------------
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="2F5496")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )
        number_fmt = '#,##0'
        date_fmt = 'DD.MM.YYYY'

        # ---- Headers ----------------------------------------------------
        headers = [
            ("ZC", 10),
            ("Klient", 28),
            ("Region", 16),
            ("Znacky", 24),
            ("Zacatek cile", 14),
            ("Konec cile", 14),
            ("Zaklad", 14),
            ("Cil", 14),
            ("Obrat celkem (netto)", 20),
        ]
        for name in BRAND_NAMES:
            headers.append((f"Obrat {name} (netto)", 20))

        headers += [
            ("Dobropisy celkem", 18),
        ]
        for name in BRAND_NAMES:
            headers.append((f"Dobropisy {name}", 18))

        headers += [
            ("Body za faktury", 16),
            ("Body za dobropisy", 18),
            ("Body za cile", 16),
        ]

        for col_idx, (title, width) in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze top row so it stays visible when scrolling
        ws.freeze_panes = "A2"

        # ---- Data rows --------------------------------------------------
        for row_idx, data in enumerate(rows, start=2):
            values = [
                data["client_number"],
                data["client_name"],
                data["region"],
                data["brands"],
                data["goal_from"],
                data["goal_to"],
                data["goal_base"],
                data["goal_value"],
                data["turnover_total"],
            ]
            for name in BRAND_NAMES:
                values.append(data[f"turnover_{name}"])

            values.append(data["credit_total"])

            for name in BRAND_NAMES:
                values.append(data[f"credit_{name}"])

            values += [
                data["standard_points"],
                data["credit_note_points"],
                data["extra_points"],
            ]

            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border

                # Apply number formatting to the appropriate columns
                # Columns 5-6 are dates
                if col_idx in (5, 6) and isinstance(val, date):
                    cell.number_format = date_fmt
                # Columns 7-8 are goal base / goal value (integers)
                elif col_idx in (7, 8):
                    cell.number_format = number_fmt
                # Columns 9 onward are numeric (turnover, credits, points)
                elif col_idx >= 9:
                    cell.number_format = number_fmt

        # ---- Auto-filter on header row ----------------------------------
        last_col_letter = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col_letter}{len(rows) + 1}"

        wb.save(filename)
