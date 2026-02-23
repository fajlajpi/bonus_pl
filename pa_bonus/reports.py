"""
Reports Framework for the Bonus System.

This module provides a pluggable report generation system. Each report is a subclass
of BaseReport that defines its own column headers and data-gathering logic. The framework
handles all shared concerns: Excel file creation, header styling, column auto-sizing,
and HTTP response packaging.

To add a new report:
    1. Create a subclass of BaseReport
    2. Define the class attributes (report_id, title, description, filename_prefix)
    3. Implement get_headers() and get_rows()
    4. The report will be automatically registered and appear on the Reports Hub page.

Architecture Notes:
    - Reports are auto-registered via __init_subclass__, so there is no manual registry.
    - All reports produce .xlsx files using openpyxl for consistency.
    - The get_rows() method should yield lists of values matching the header order.
      This is a deliberate design choice over returning dicts, because it avoids the
      overhead of dict key lookups for every cell in potentially large exports.
"""

import io
import logging
from abc import abstractmethod
from datetime import date

from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Sum, Q, Value, DecimalField
from django.db.models.functions import Coalesce

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Report Registry
# ---------------------------------------------------------------------------

_report_registry: dict[str, type["BaseReport"]] = {}


def get_all_reports() -> list[type["BaseReport"]]:
    """Return all registered report classes, sorted by their display order."""
    return sorted(_report_registry.values(), key=lambda r: r.display_order)


def get_report_by_id(report_id: str) -> type["BaseReport"] | None:
    """Look up a report class by its unique identifier."""
    return _report_registry.get(report_id)


# ---------------------------------------------------------------------------
# Shared Styles
# ---------------------------------------------------------------------------

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_BORDER = Border(
    bottom=Side(style="thin", color="2F5597"),
)

DATA_FONT = Font(name="Calibri", size=11)
DATA_ALIGNMENT = Alignment(vertical="center")

# These can be used by subclasses for special columns (e.g. currency, percentages)
NUMBER_FORMAT_PERCENT = "0.00%"
NUMBER_FORMAT_CURRENCY = '#,##0.00'
NUMBER_FORMAT_INTEGER = '#,##0'


# ---------------------------------------------------------------------------
# Base Report
# ---------------------------------------------------------------------------

class BaseReport:
    """
    Abstract base class for all Excel reports.

    Subclasses MUST define:
        report_id (str):        Unique slug used in URLs and form values.
        title (str):            Human-readable name shown on the Reports Hub.
        description (str):      Short explanation of what the report contains.
        filename_prefix (str):  Prefix for the downloaded file name.
        display_order (int):    Order in which the report appears on the hub page.

    Subclasses MUST implement:
        get_headers() -> list[str]
        get_rows() -> Iterable[list]

    Subclasses MAY override:
        get_column_formats() -> dict[int, str]
            Return a mapping of zero-based column index to openpyxl number format
            strings.  For example, {7: NUMBER_FORMAT_PERCENT} to format column H
            as a percentage.
        get_sheet_name() -> str
            Return a custom sheet name (default is the report title, truncated to 31 chars).
    """

    # -- Subclass attributes (must be overridden) --
    report_id: str = ""
    title: str = ""
    description: str = ""
    filename_prefix: str = "report"
    display_order: int = 100

    def __init_subclass__(cls, **kwargs):
        """Auto-register every concrete subclass that defines a report_id."""
        super().__init_subclass__(**kwargs)
        if cls.report_id:
            _report_registry[cls.report_id] = cls

    # -- Abstract interface --

    @abstractmethod
    def get_headers(self) -> list[str]:
        """Return an ordered list of column header strings."""
        ...

    @abstractmethod
    def get_rows(self) -> list[list]:
        """
        Return an iterable of rows where each row is a list of values
        matching the order of get_headers().
        """
        ...

    # -- Optional overrides --

    def get_column_formats(self) -> dict[int, str]:
        """
        Return a dict mapping zero-based column index to an openpyxl number
        format string.  Override this in subclasses when you need specific
        formatting for numeric columns.
        """
        return {}

    def get_sheet_name(self) -> str:
        """Return the worksheet tab name (max 31 chars per Excel spec)."""
        return self.title[:31]

    # -- Core generation logic (not intended to be overridden) --

    def generate_workbook(self) -> Workbook:
        """
        Build and return a fully-populated openpyxl Workbook.

        This is the workhorse method.  It:
            1. Creates the workbook and active sheet
            2. Writes styled headers
            3. Writes all data rows
            4. Applies number formats
            5. Auto-sizes columns
        """
        wb = Workbook()
        ws = wb.active
        ws.title = self.get_sheet_name()

        headers = self.get_headers()
        column_formats = self.get_column_formats()

        # -- Write header row --
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = HEADER_BORDER

        # Freeze the header row so it stays visible when scrolling
        ws.freeze_panes = "A2"

        # -- Write data rows --
        for row_idx, row_data in enumerate(self.get_rows(), start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = DATA_FONT
                cell.alignment = DATA_ALIGNMENT

                # Apply column-specific number format if defined
                fmt = column_formats.get(col_idx - 1)
                if fmt:
                    cell.number_format = fmt

        # -- Auto-size columns --
        self._auto_size_columns(ws, headers)

        return wb

    def generate_response(self) -> HttpResponse:
        """
        Generate the workbook and wrap it in an HttpResponse with the correct
        content type and Content-Disposition for browser download.
        """
        wb = self.generate_workbook()

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        timestamp = timezone.now().strftime("%Y%m%d_%H%M")
        filename = f"{self.filename_prefix}_{timestamp}.xlsx"

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    # -- Private helpers --

    @staticmethod
    def _auto_size_columns(ws, headers: list[str]):
        """
        Set each column width to the maximum of the header length and the
        longest data value in that column, with a reasonable cap.

        This is an approximation -- openpyxl does not support true auto-fit
        because that is a client-side feature in Excel.  The heuristic here
        works well for typical data.
        """
        MIN_WIDTH = 10
        MAX_WIDTH = 50

        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = len(str(headers[col_idx - 1]))

            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))

            # Apply width with padding, clamped to our bounds
            adjusted = min(max(max_length + 3, MIN_WIDTH), MAX_WIDTH)
            ws.column_dimensions[col_letter].width = adjusted


# ---------------------------------------------------------------------------
# Concrete Reports
# ---------------------------------------------------------------------------


class AllClientsReport(BaseReport):
    """
    Report 1: Full client export.

    Lists every user in the system (excluding staff/superusers) with their
    contract status, available and incoming points, and current extra goal
    data if applicable.
    """

    report_id = "all_clients"
    title = "All Clients"
    description = (
        "Complete client listing with contract status, available/incoming points, "
        "and current extra goal progress.  Export to Excel for ad-hoc filtering."
    )
    filename_prefix = "all_clients_report"
    display_order = 10

    def get_headers(self) -> list[str]:
        return [
            "Client Number",
            "Last Name",
            "First Name",
            "Email",
            "Region",
            "Sales Rep",
            "Is Active",
            "Has Active Contract",
            "Contract From",
            "Contract To",
            "Available Points",
            "Incoming Points",
            "Has Extra Goal",
            "Goal Value",
            "Goal Base",
            "Goal Period From",
            "Goal Period To",
            "Goal Brands",
            "Current Turnover",
            "Goal Percentage",
        ]

    def get_column_formats(self) -> dict[int, str]:
        return {
            18: NUMBER_FORMAT_CURRENCY,   # Current Turnover
            19: NUMBER_FORMAT_PERCENT,    # Goal Percentage
        }

    def get_rows(self) -> list[list]:
        from pa_bonus.models import User, UserContract, UserContractGoal, PointsTransaction
        from pa_bonus.utilities import calculate_turnover_for_goal

        today = timezone.now().date()

        # Fetch all non-staff users with annotated point balances in a single query.
        # This avoids N+1 queries -- one of the most common performance pitfalls.
        users = (
            User.objects
            .filter(is_staff=False, is_superuser=False)
            .select_related("region")
            .annotate(
                available_points=Coalesce(
                    Sum("pointstransaction__value",
                        filter=Q(pointstransaction__status="CONFIRMED")),
                    Value(0),
                ),
                incoming_points=Coalesce(
                    Sum("pointstransaction__value",
                        filter=Q(pointstransaction__status="PENDING")),
                    Value(0),
                ),
            )
            .order_by("last_name", "first_name")
        )

        rows = []
        for user in users:
            # Active contract
            contract = (
                UserContract.objects
                .filter(user_id=user, is_active=True)
                .first()
            )

            # Current extra goal -- the one whose period covers today
            current_goal = None
            goal_turnover = None
            goal_percentage = None
            if contract:
                current_goal = (
                    UserContractGoal.objects
                    .filter(
                        user_contract=contract,
                        goal_period_from__lte=today,
                        goal_period_to__gte=today,
                    )
                    .prefetch_related("brands")
                    .first()
                )
                if current_goal:
                    goal_turnover = float(calculate_turnover_for_goal(
                        user,
                        current_goal.brands.all(),
                        current_goal.goal_period_from,
                        min(today, current_goal.goal_period_to),
                    ))
                    goal_percentage = (
                        goal_turnover / current_goal.goal_value
                        if current_goal.goal_value > 0 else 0
                    )

            sales_rep = user.get_sales_rep()

            rows.append([
                user.user_number,
                user.last_name,
                user.first_name,
                user.email,
                user.region.name if user.region else "",
                sales_rep.get_full_name() if sales_rep else "",
                "Yes" if user.is_active else "No",
                "Yes" if contract else "No",
                contract.contract_date_from if contract else "",
                contract.contract_date_to if contract else "",
                user.available_points,
                user.incoming_points,
                "Yes" if current_goal else "No",
                current_goal.goal_value if current_goal else "",
                current_goal.goal_base if current_goal else "",
                current_goal.goal_period_from if current_goal else "",
                current_goal.goal_period_to if current_goal else "",
                ", ".join(b.name for b in current_goal.brands.all()) if current_goal else "",
                goal_turnover if goal_turnover is not None else "",
                goal_percentage if goal_percentage is not None else "",
            ])

        return rows


class ExtraGoalsReport(BaseReport):
    """
    Report 2: Extra goals progress.

    Lists all clients who have a currently-running extra goal, together with
    their total-period progress (not the current milestone period).
    """

    report_id = "extra_goals"
    title = "Extra Goals Progress"
    description = (
        "All clients with a current extra goal and their progress towards it "
        "over the full goal period."
    )
    filename_prefix = "extra_goals_report"
    display_order = 20

    def get_headers(self) -> list[str]:
        return [
            "Client Number",
            "Last Name",
            "First Name",
            "Region",
            "Sales Rep",
            "Goal Brands",
            "Goal Period From",
            "Goal Period To",
            "Goal Value",
            "Goal Base",
            "Current Turnover",
            "Turnover vs Goal",
            "Goal Percentage",
            "Turnover Remaining",
            "Evaluation Frequency (months)",
            "Bonus Percentage",
        ]

    def get_column_formats(self) -> dict[int, str]:
        return {
            10: NUMBER_FORMAT_CURRENCY,   # Current Turnover
            11: NUMBER_FORMAT_CURRENCY,   # Turnover vs Goal
            12: NUMBER_FORMAT_PERCENT,    # Goal Percentage
            13: NUMBER_FORMAT_CURRENCY,   # Turnover Remaining
            15: NUMBER_FORMAT_PERCENT,    # Bonus Percentage
        }

    def get_rows(self) -> list[list]:
        from pa_bonus.models import UserContractGoal
        from pa_bonus.utilities import calculate_turnover_for_goal

        today = timezone.now().date()

        goals = (
            UserContractGoal.objects
            .filter(goal_period_from__lte=today, goal_period_to__gte=today)
            .select_related("user_contract__user_id__region")
            .prefetch_related("brands")
            .order_by("user_contract__user_id__last_name")
        )

        rows = []
        for goal in goals:
            user = goal.user_contract.user_id
            turnover = float(calculate_turnover_for_goal(
                user,
                goal.brands.all(),
                goal.goal_period_from,
                min(today, goal.goal_period_to),
            ))
            percentage = turnover / goal.goal_value if goal.goal_value > 0 else 0
            remaining = max(0, goal.goal_value - turnover)
            sales_rep = user.get_sales_rep()

            rows.append([
                user.user_number,
                user.last_name,
                user.first_name,
                user.region.name if user.region else "",
                sales_rep.get_full_name() if sales_rep else "",
                ", ".join(b.name for b in goal.brands.all()),
                goal.goal_period_from,
                goal.goal_period_to,
                goal.goal_value,
                goal.goal_base,
                turnover,
                turnover - goal.goal_value,
                percentage,
                remaining,
                goal.evaluation_frequency,
                goal.bonus_percentage,
            ])

        return rows


class PointsReport(BaseReport):
    """
    Report 3: Points overview.

    Lists all clients with their available (confirmed) and incoming (pending)
    point balances.
    """

    report_id = "points"
    title = "Points Overview"
    description = (
        "All clients with their current available (confirmed) and incoming "
        "(pending) point balances."
    )
    filename_prefix = "points_report"
    display_order = 30

    def get_headers(self) -> list[str]:
        return [
            "Client Number",
            "Last Name",
            "First Name",
            "Email",
            "Region",
            "Sales Rep",
            "Is Active",
            "Available Points",
            "Incoming Points",
            "Total Points",
        ]

    def get_column_formats(self) -> dict[int, str]:
        return {
            7: NUMBER_FORMAT_INTEGER,
            8: NUMBER_FORMAT_INTEGER,
            9: NUMBER_FORMAT_INTEGER,
        }

    def get_rows(self) -> list[list]:
        from pa_bonus.models import User

        users = (
            User.objects
            .filter(is_staff=False, is_superuser=False)
            .select_related("region")
            .annotate(
                available_points=Coalesce(
                    Sum("pointstransaction__value",
                        filter=Q(pointstransaction__status="CONFIRMED")),
                    Value(0),
                ),
                incoming_points=Coalesce(
                    Sum("pointstransaction__value",
                        filter=Q(pointstransaction__status="PENDING")),
                    Value(0),
                ),
            )
            .order_by("last_name", "first_name")
        )

        rows = []
        for user in users:
            sales_rep = user.get_sales_rep()
            rows.append([
                user.user_number,
                user.last_name,
                user.first_name,
                user.email,
                user.region.name if user.region else "",
                sales_rep.get_full_name() if sales_rep else "",
                "Yes" if user.is_active else "No",
                user.available_points,
                user.incoming_points,
                user.available_points + user.incoming_points,
            ])

        return rows


class RewardRequestsReport(BaseReport):
    """
    Report 4: Reward requests summary.

    Lists all reward requests with status and point values, without
    itemising the individual rewards.
    """

    report_id = "reward_requests"
    title = "Reward Requests"
    description = (
        "All reward requests with their status, point value, and notes. "
        "No itemised reward listing."
    )
    filename_prefix = "reward_requests_report"
    display_order = 40

    def get_headers(self) -> list[str]:
        return [
            "Request ID",
            "Client Number",
            "Last Name",
            "First Name",
            "Region",
            "Sales Rep",
            "Requested At",
            "Status",
            "Total Points",
            "Customer Note",
            "Manager Note",
        ]

    def get_column_formats(self) -> dict[int, str]:
        return {
            8: NUMBER_FORMAT_INTEGER,
        }

    def get_rows(self) -> list[list]:
        from pa_bonus.models import RewardRequest

        requests = (
            RewardRequest.objects
            .select_related("user__region")
            .order_by("-requested_at")
        )

        rows = []
        for req in requests:
            user = req.user
            sales_rep = user.get_sales_rep()
            rows.append([
                req.id,
                user.user_number,
                user.last_name,
                user.first_name,
                user.region.name if user.region else "",
                sales_rep.get_full_name() if sales_rep else "",
                req.requested_at.strftime("%Y-%m-%d %H:%M"),
                req.get_status_display(),
                req.total_points,
                req.note or "",
                req.description or "",
            ])

        return rows


class ItemisedRewardsReport(BaseReport):
    """
    Report 5: Itemised rewards for pending and accepted requests.

    One row per reward-request-item, including the reward's current stock
    availability so you can compare against live stock levels.
    """

    report_id = "itemised_rewards"
    title = "Itemised Reward Requests"
    description = (
        "All individual reward items from pending and accepted requests, with "
        "quantities, point costs, and current stock availability levels."
    )
    filename_prefix = "itemised_rewards_report"
    display_order = 50

    def get_headers(self) -> list[str]:
        return [
            "Request ID",
            "Request Date",
            "Request Status",
            "Client Number",
            "Last Name",
            "First Name",
            "Sales Rep",
            "Reward Code",
            "Reward Name",
            "Quantity",
            "Point Cost (per unit)",
            "Total Point Cost",
            "Current Stock Status",
        ]

    def get_column_formats(self) -> dict[int, str]:
        return {
            9: NUMBER_FORMAT_INTEGER,
            10: NUMBER_FORMAT_INTEGER,
            11: NUMBER_FORMAT_INTEGER,
        }

    def get_rows(self) -> list[list]:
        from pa_bonus.models import RewardRequestItem

        items = (
            RewardRequestItem.objects
            .filter(reward_request__status__in=["PENDING", "ACCEPTED"])
            .select_related(
                "reward_request__user__region",
                "reward",
            )
            .order_by("-reward_request__requested_at", "reward__abra_code")
        )

        rows = []
        for item in items:
            req = item.reward_request
            user = req.user
            reward = item.reward
            sales_rep = user.get_sales_rep()

            rows.append([
                req.id,
                req.requested_at.strftime("%Y-%m-%d %H:%M"),
                req.get_status_display(),
                user.user_number,
                user.last_name,
                user.first_name,
                sales_rep.get_full_name() if sales_rep else "",
                reward.abra_code,
                reward.name,
                item.quantity,
                item.point_cost,
                item.quantity * item.point_cost,
                reward.get_availability_display(),
            ])

        return rows


class PreviousExtraGoalsReport(BaseReport):
    """
    Report 6: Completed (historical) extra goals.

    Lists all extra goals whose period has already ended, together with the
    final turnover figures and the evaluation outcomes.  This is the retrospective
    counterpart to ExtraGoalsReport, which only covers currently-running goals.

    For each completed goal, the report shows:
        - The same client and goal metadata as the active goals report
        - Final turnover for the full goal period
        - Whether the overall goal target was reached
        - Whether the goal has been formally evaluated by a manager
        - How many milestone periods were evaluated and how many were achieved
        - Total bonus points awarded across all evaluations
    """

    report_id = "previous_extra_goals"
    title = "Previous Extra Goals (Completed)"
    description = (
        "All extra goals whose period has ended.  Shows the final turnover, "
        "whether the target was met, evaluation status, and total bonus points awarded."
    )
    filename_prefix = "previous_extra_goals_report"
    display_order = 25  # Slots right after the active extra goals report

    def get_headers(self) -> list[str]:
        return [
            "Client Number",
            "Last Name",
            "First Name",
            "Region",
            "Sales Rep",
            "Goal Brands",
            "Goal Period From",
            "Goal Period To",
            "Goal Value",
            "Goal Base",
            "Final Turnover",
            "Turnover vs Goal",
            "Goal Percentage",
            "Goal Target Met",
            "Evaluation Status",
            "Milestones Evaluated",
            "Milestones Achieved",
            "Total Points Awarded",
            "Evaluation Frequency (months)",
            "Bonus Percentage",
        ]

    def get_column_formats(self) -> dict[int, str]:
        return {
            10: NUMBER_FORMAT_CURRENCY,   # Final Turnover
            11: NUMBER_FORMAT_CURRENCY,   # Turnover vs Goal
            12: NUMBER_FORMAT_PERCENT,    # Goal Percentage
            17: NUMBER_FORMAT_INTEGER,    # Total Points Awarded
            19: NUMBER_FORMAT_PERCENT,    # Bonus Percentage
        }

    def get_rows(self) -> list[list]:
        from pa_bonus.models import UserContractGoal, GoalEvaluation
        from pa_bonus.utilities import calculate_turnover_for_goal

        today = timezone.now().date()

        # Completed goals: the period end date is strictly in the past.
        goals = (
            UserContractGoal.objects
            .filter(goal_period_to__lt=today)
            .select_related("user_contract__user_id__region")
            .prefetch_related("brands", "evaluations")
            .order_by(
                "user_contract__user_id__last_name",
                "-goal_period_to",
            )
        )

        rows = []
        for goal in goals:
            user = goal.user_contract.user_id

            # Full-period turnover.  Because the period has ended we use the
            # actual goal_period_to rather than min(today, ...).
            final_turnover = float(calculate_turnover_for_goal(
                user,
                goal.brands.all(),
                goal.goal_period_from,
                goal.goal_period_to,
            ))

            percentage = (
                final_turnover / goal.goal_value if goal.goal_value > 0 else 0
            )
            target_met = final_turnover >= goal.goal_value

            # ----------------------------------------------------------
            # Evaluation data
            # ----------------------------------------------------------
            # All evaluation records for this goal (prefetched above).
            evaluations = list(goal.evaluations.all())

            # How many evaluation periods *should* exist?
            expected_periods = goal.get_evaluation_periods()
            total_expected = len(expected_periods)

            total_evaluated = len(evaluations)
            milestones_achieved = sum(1 for e in evaluations if e.is_achieved)
            total_points_awarded = sum(e.bonus_points for e in evaluations)

            # Determine a human-readable evaluation status.
            # "Fully Evaluated"  -- every expected period has a record.
            # "Partially Evaluated" -- some records exist but not all.
            # "Not Evaluated"    -- no evaluation records at all.
            if total_evaluated == 0:
                eval_status = "Not Evaluated"
            elif total_evaluated >= total_expected:
                eval_status = "Fully Evaluated"
            else:
                eval_status = "Partially Evaluated"

            sales_rep = user.get_sales_rep()

            rows.append([
                user.user_number,
                user.last_name,
                user.first_name,
                user.region.name if user.region else "",
                sales_rep.get_full_name() if sales_rep else "",
                ", ".join(b.name for b in goal.brands.all()),
                goal.goal_period_from,
                goal.goal_period_to,
                goal.goal_value,
                goal.goal_base,
                final_turnover,
                final_turnover - goal.goal_value,
                percentage,
                "Yes" if target_met else "No",
                eval_status,
                f"{total_evaluated} / {total_expected}",
                f"{milestones_achieved} / {total_evaluated}" if total_evaluated > 0 else "N/A",
                total_points_awarded,
                goal.evaluation_frequency,
                goal.bonus_percentage,
            ])

        return rows
