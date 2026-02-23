from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import permission_required
from django.contrib import messages
from django.views.generic import ListView, View
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.core.paginator import Paginator
from django.utils import timezone
from django.db.models import Sum, Count, Q, F, Case, When, IntegerField
from django.db import transaction
from django.urls import reverse
import logging
from pa_bonus.forms import FileUploadForm, ClientCreationForm
from pa_bonus.tasks import process_uploaded_file, process_stock_file
from pa_bonus.models import (FileUpload, Reward, RewardRequest, RewardRequestItem, PointsTransaction,
                             EmailNotification, User, Region, UserContract, InvoiceBrandTurnover, Brand,
                             UserActivity, UserContractGoal, GoalEvaluation)
from pa_bonus.utilities import ManagerGroupRequiredMixin, calculate_turnover_for_goal

from pa_bonus.exports import generate_telemarketing_export

from datetime import datetime, timedelta, date
from dateutil.relativedelta import relativedelta

import openpyxl
from openpyxl.styles import Font, PatternFill
import io


logger = logging.getLogger(__name__)

class ManagerDashboardView(ManagerGroupRequiredMixin, View):
    """
    Main dashboard view for managers.
    
    Provides an overview of system status and links to manager functions.
    """
    template_name = 'manager/dashboard.html'
    
    def get(self, request):
        # System-wide points statistics
        from django.db.models import Sum, Count, Q, F, Value
        from django.db.models.functions import Coalesce
        
        # 1. Total points summary
        points_summary = PointsTransaction.objects.filter(
            status__in=['PENDING', 'CONFIRMED']
        ).values('status').annotate(
            total=Coalesce(Sum('value'), Value(0))
        ).order_by('status')
        
        # Convert to a dictionary for easier access in template
        points_data = {
            'PENDING': 0,
            'CONFIRMED': 0,
        }
        for entry in points_summary:
            points_data[entry['status']] = entry['total']
            
        # 2. Reward requests statistics
        request_stats = RewardRequest.objects.filter(
            status__in=['PENDING', 'ACCEPTED']
        ).values('status').annotate(
            count=Count('id'),
            total_points=Coalesce(Sum('total_points'), Value(0))
        ).order_by('status')
        
        # Convert to dictionary
        request_data = {
            'PENDING': {'count': 0, 'total_points': 0},
            'ACCEPTED': {'count': 0, 'total_points': 0},
        }
        for entry in request_stats:
            request_data[entry['status']] = {
                'count': entry['count'],
                'total_points': entry['total_points']
            }
            
        # 3. Top 10 clients by available points
        top_clients = User.objects.annotate(
            available_points=Coalesce(
                Sum('pointstransaction__value', 
                    filter=Q(pointstransaction__status='CONFIRMED')),
                Value(0)
            ),
            pending_points=Coalesce(
                Sum('pointstransaction__value', 
                    filter=Q(pointstransaction__status='PENDING')),
                Value(0)
            )
        ).filter(
            available_points__gt=0
        ).order_by('-available_points')[:10]

        # NEW: Add goal evaluation statistics
        today = timezone.now().date()
        
        # Count pending goal evaluations
        pending_evaluations = 0
        total_potential_points = 0
        
        # Get all active goals
        active_goals = UserContractGoal.objects.filter(
            goal_period_from__lte=today
        ).select_related('user_contract__user_id')
        
        for goal in active_goals:
            periods = goal.get_evaluation_periods()
            for start, end, is_final in periods:
                if end < today:  # Period has ended
                    # Check if already evaluated
                    if not goal.evaluations.filter(period_start=start, period_end=end).exists():
                        pending_evaluations += 1
                        
                        # Calculate potential points (rough estimate)
                        targets = goal.get_period_targets(start, end)
                        actual = calculate_turnover_for_goal(
                            goal.user_contract.user_id,
                            goal.brands.all(),
                            start, end
                        )
                        if actual > targets['goal_value']:
                            potential_points = int((float(actual) - targets['goal_base']) * goal.bonus_percentage)
                            total_potential_points += max(0, potential_points)
        
        context = {
            'points_data': points_data,
            'request_data': request_data,
            'top_clients': top_clients,
            'goal_stats': {
                'pending_evaluations': pending_evaluations,
                'potential_points': total_potential_points
                }
        }
        
        return render(request, self.template_name, context)
    
    
@permission_required('pa_bonus.add_fileupload', raise_exception=True)
def upload_file(request):
    """
    Handles file uploads for processing invoice data.

    This view allows users with the correct permission to upload invoice data files.
    After a successful upload, the file is processed accordingly.

    Args:
        request (HttpRequest): The HTTP request object containing the file upload.

    Returns:
        HttpResponse: Renders the upload form (GET) or redirects to the upload history (POST).
    """
    if request.method == "POST":
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            upload = form.save(commit=False)
            upload.uploaded_by = request.user
            upload.save()

            # Process the uploaded file
            try:
                process_uploaded_file(upload.id)
                messages.success(
                    request, 
                    'File uploaded successfully and is being processed'
                )
            except Exception as e:
                messages.error(
                    request, 
                    f'Error processing file: {str(e)}'
                )

            return redirect('upload_history')
    else:
        form = FileUploadForm()
    
    return render(request, 'upload.html', {'form': form})

@permission_required('pa_bonus.change_reward', raise_exception=True)
def upload_stock(request):
    """
    Handles file uploads for processing stock data and updating reward availability.

    This view allows users with the correct permission to upload stock data files.
    After a successful upload, the file is processed to update reward availability status.

    Args:
        request (HttpRequest): The HTTP request object containing the file upload.

    Returns:
        HttpResponse: Renders the upload form (GET) or redirects to the reward list (POST).
    """
    if request.method == "POST":
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            upload = form.save(commit=False)
            upload.uploaded_by = request.user
            upload.save()

            # Process the uploaded file
            try:
                process_stock_file(upload.id)
                messages.success(
                    request, 
                    'Stock file uploaded successfully and rewards have been updated'
                )
            except Exception as e:
                messages.error(
                    request, 
                    f'Error processing file: {str(e)}'
                )

            return redirect('upload_history')
    else:
        form = FileUploadForm()
    
    return render(request, 'manager/upload_stock.html', {'form': form})

@permission_required('pa_bonus.add_fileupload', raise_exception=True)
def upload_history(request):
    """
    Displays the history of uploaded files.

    This view lists all uploaded files in the order of uploading.

    Args:
        request (HttpRequest): The HTTP request object.

    Returns:
        HttpResponse: Renders the upload history template with the list of uploads.
    """
    uploads = FileUpload.objects.all().order_by('-uploaded_at')
    return render(request, 'upload_history.html', {'uploads': uploads})

class ManagerRewardRequestListView(ManagerGroupRequiredMixin, ListView):
    """
    (Managers Only) Lists the current reward requests in the system.

    Attributes:
        template_name (str): Name of template to render
        context_object_name (str): Name of the context object we're working with in the ListView
        paginate_by (int): Number of requests per page
    """
    template_name = 'manager/reward_requests_list.html'
    context_object_name = 'reward_requests'
    # paginate_by = 25

    def get_queryset(self):
        """
        Get the data (queryset) to populate the ListView with. With filtering by 'status'.
        """
        queryset = RewardRequest.objects.select_related('user').order_by('-requested_at')
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        return queryset
    
class ManagerRewardRequestDetailView(ManagerGroupRequiredMixin, View):
    """
    (Managers Only) Detail of Reward Request with editing and confirming capability.
    Updated to show client information and read-only item list.
    """
    template_name = 'manager/reward_request_detail.html'

    def get(self, request, pk):
        reward_request = get_object_or_404(RewardRequest, pk=pk)
        items = reward_request.rewardrequestitem_set.select_related('reward').filter(quantity__gt=0)
        user_balance = reward_request.user.get_balance()
        
        return render(request, self.template_name, {
            'request_obj': reward_request,
            'items': items,
            'user_balance': user_balance,
        })

    @transaction.atomic
    def post(self, request, pk):
        reward_request = get_object_or_404(RewardRequest, pk=pk)
        old_status = reward_request.status
        
        # Update the customer note
        customer_note = request.POST.get('customer_note', '')
        reward_request.note = customer_note
        
        # Update the request status and description
        new_status = request.POST.get('status')
        reward_request.description = request.POST.get('manager_message', '')
        reward_request.status = new_status
        reward_request.save()
        
        # Update the point transaction to match the current state
        self._update_point_transaction(reward_request, old_status, new_status)
        
        messages.success(request, f"Request {reward_request.pk} updated.")
        return redirect('manager_reward_requests')
    
    def _update_point_transaction(self, reward_request, old_status, new_status):
        """Update the point transaction to match the current state of the request."""
        transaction = self._get_reward_transaction(reward_request)
        if not transaction:
            return
        
        # If request is rejected/cancelled, cancel the transaction
        if new_status in ['REJECTED', 'CANCELLED']:
            transaction.status = 'CANCELLED'
            transaction.save()
        
        # If request was rejected/cancelled but is now active, reactivate transaction
        elif old_status in ['REJECTED', 'CANCELLED'] and new_status not in ['REJECTED', 'CANCELLED']:
            transaction.status = 'CONFIRMED'
            transaction.save()
        
        # In all cases, ensure the transaction amount matches the request total
        if transaction.status == 'CONFIRMED':
            transaction.value = -reward_request.total_points
            transaction.save()
    
    def _get_reward_transaction(self, reward_request):
        """Get the associated reward claim transaction."""
        try:
            return PointsTransaction.objects.get(
                reward_request=reward_request,
                type='REWARD_CLAIM'
            )
        except PointsTransaction.DoesNotExist:
            logger.warning(f"No transaction found for reward request {reward_request.id}")
            return None
        except PointsTransaction.MultipleObjectsReturned:
            logger.error(f"Multiple transactions found for reward request {reward_request.id}")
            messages.warning(self.request, "Multiple transactions found for this request. Please check manually.")
            return None

        
class ExportTelemarketingFileView(ManagerGroupRequiredMixin, View):
    """
    Export a telemarketing file for a specific reward request
    """
    def get(self, request, pk):
        output = generate_telemarketing_export(pk)
        
        if output is None:
            messages.error(request, "Reward request not found or not in ACCEPTED status.")
            return redirect('manager_reward_requests')
        
        # Prepare response
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=reward_request_{pk}_{timezone.now().strftime("%Y%m%d")}.xlsx'
        
        messages.success(request, f"Reward request {pk} has been exported and marked as FINISHED.")
        return response


class TransactionApprovalView(ManagerGroupRequiredMixin, View):
    """
    View for managers to approve pending transactions based on month/year.
    
    Allows managers to see and approve transactions that are due for approval
    (those from three months ago) in a simple interface.
    """
    template_name = 'manager/transaction_approval.html'
    
    def get(self, request):
        """
        Display the approval form and optionally show transactions for a selected month.
        """
        today = timezone.now().date()
        
        # Default to showing transactions from 3 months ago
        default_approval_date = today - relativedelta(months=3)
        default_year = default_approval_date.year
        default_month = default_approval_date.month
        
        # Get user-selected month and year if provided
        selected_year = int(request.GET.get('year', default_year))
        selected_month = int(request.GET.get('month', default_month))
        
        # Generate a list of years (from 2 years ago to current year)
        available_years = range(today.year - 2, today.year + 1)
        
        # Get month range for filtering
        start_date = date(selected_year, selected_month, 1)
        if selected_month == 12:
            end_date = date(selected_year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(selected_year, selected_month + 1, 1) - timedelta(days=1)
        
        # Get pending transactions for the selected month
        pending_transactions = PointsTransaction.objects.filter(
            status='PENDING',
            date__gte=start_date,
            date__lte=end_date
        ).select_related('user', 'brand')
        
        # Get statistics for the selected month
        stats = pending_transactions.aggregate(
            total_transactions=Count('id'),
            total_points=Sum('value')
        )
        
        # Determine if we should highlight this month for approval
        # (if it's the month that is due for approval based on the 3-month rule)
        is_approval_month = (
            selected_year == default_approval_date.year and 
            selected_month == default_approval_date.month
        )
        
        # Get available months (1-12)
        available_months = [(i, date(2000, i, 1).strftime('%B')) for i in range(1, 13)]
        
        context = {
            'pending_transactions': pending_transactions,
            'stats': stats,
            'selected_year': selected_year,
            'selected_month': selected_month,
            'month_name': date(selected_year, selected_month, 1).strftime('%B'),
            'available_years': available_years,
            'available_months': available_months,
            'is_approval_month': is_approval_month,
            'start_date': start_date,
            'end_date': end_date,
        }
        
        return render(request, self.template_name, context)
    
    @transaction.atomic
    def post(self, request):
        """
        Process the approval of transactions for the selected month.
        """
        selected_year = int(request.POST.get('year'))
        selected_month = int(request.POST.get('month'))
        
        # Get month range for filtering
        start_date = date(selected_year, selected_month, 1)
        if selected_month == 12:
            end_date = date(selected_year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(selected_year, selected_month + 1, 1) - timedelta(days=1)
        
        # Update pending transactions to confirmed
        pending_transactions = PointsTransaction.objects.filter(
            status='PENDING',
            date__gte=start_date,
            date__lte=end_date
        )
        
        # Count before updating for the message
        transaction_count = pending_transactions.count()
        points_total = pending_transactions.aggregate(total=Sum('value'))['total'] or 0
        
        # Update the transactions
        pending_transactions.update(status='CONFIRMED')
        
        # Schedule email notifications for each user with confirmed transactions
        self.schedule_email_notifications(pending_transactions)
        
        # Success message
        messages.success(
            request, 
            f"Successfully approved {transaction_count} transactions totaling {points_total} points."
        )
        
        # Redirect back to the form
        return redirect('transaction_approval')
    
    def schedule_email_notifications(self, transactions):
        """
        Schedule email notifications for users whose transactions were approved.
        
        Creates EmailNotification records for each user who had transactions approved.
        These will be processed by a separate task/process.
        
        Args:
            transactions: QuerySet of approved transactions
        """
        # Get unique users who had transactions approved
        user_ids = transactions.values_list('user_id', flat=True).distinct()
        
        # For each user, create a notification
        for user_id in user_ids:
            # Get the user's transactions that were just approved
            user_transactions = transactions.filter(user_id=user_id)
            user = user_transactions.first().user
            
            # Calculate total points
            total_points = user_transactions.aggregate(total=Sum('value'))['total'] or 0
            
            # Create notification message
            subject = "Your bonus points have been confirmed!"
            message = f"""
Dear {user.first_name} {user.last_name},

We are pleased to inform you that your transactions for {user_transactions.first().date.strftime('%B %Y')} 
have been confirmed, adding {total_points} points to your account.

Your current point balance is now: {user.get_balance()} points.

You can log in to the Bonus Program portal to view these transactions and explore 
available rewards.

Thank you for your business!

Best regards,
The Bonus Program Team
            """
            
            # Create the notification record
            EmailNotification.objects.create(
                user=user,
                subject=subject,
                message=message,
                status='PENDING'
            )


class SMSExportView(ManagerGroupRequiredMixin, View):
    """
    Generates a CSV file for SMS notifications to clients.
    
    This view allows managers to generate a CSV file in the format required by smsbrana.cz
    to send monthly SMS notifications to clients about their point balances.
    Supports both standard and custom message templates with variable substitution.
    """
    template_name = 'manager/sms_export.html'
    
    def get(self, request):
        """
        Display the SMS export form with options.
        """
        # Get all regions for the dropdown
        regions = Region.objects.filter(is_active=True).order_by('name')
        
        context = {
            'regions': regions
        }
        
        return render(request, self.template_name, context)
    
    def post(self, request):
        """
        Generate and return the SMS export CSV file.
        """
        import csv
        from django.http import HttpResponse
        from django.utils import timezone
        
        # Create the HttpResponse object with CSV header
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="sms_export_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'
        
        # Create CSV writer with semicolon delimiter
        writer = csv.writer(response, delimiter=';')
        
        # Get active users with phone numbers
        users = User.objects.filter(is_active=True).exclude(user_phone='')
        
        # Filter by region if specified
        region_id = request.POST.get('region')
        if region_id and region_id != 'all':
            users = users.filter(region_id=region_id)
        
        # Set minimum points threshold
        min_points = request.POST.get('min_points', 0)
        try:
            min_points = int(min_points)
        except ValueError:
            min_points = 0
        
        # Determine message type and template
        message_type = request.POST.get('message_type', 'default')
        
        if message_type == 'custom':
            # Get custom message template
            message_template = request.POST.get('custom_message_text', '')
            if not message_template:
                # Fallback to default if custom template is empty
                message_template = "OS: Bonus Primavera Andorrana - na konte mate {balance} bodu. Cerpani a informace: https://bonus.primavera-and.cz/ Odhlaseni: SMS STOP na +420778799900."
        else:
            # Default message template
            message_template = "OS: Bonus Primavera Andorrana - na konte mate {balance} bodu. Cerpani a informace: https://bonus.primavera-and.cz/ Odhlaseni: SMS STOP na +420778799900."
        
        # Count for reporting
        total_sms = 0
        
        # Write SMS data rows
        for user in users:
            # Get user balance
            balance = user.get_balance()
            
            # Skip users with balance below minimum (if specified)
            if balance < min_points:
                continue
            
            # Format phone number correctly
            phone = user.user_phone.strip()
            if not phone.startswith('+'):
                # Add Czech prefix if not present
                if not phone.startswith('420'):
                    phone = '+420' + phone
                else:
                    phone = '+' + phone
            
            # Replace variables in the message template
            sms_text = message_template.format(
                balance=balance,
                first_name=user.first_name,
                last_name=user.last_name,
                user_name=user.username,
                user_number=user.user_number,
                user_email=user.email,
                user_phone=user.user_phone,
                region=user.region.name if user.region else ''
            )
            
            # Write to CSV
            writer.writerow([phone, sms_text])
            total_sms += 1
        
        # Inform user about how many SMS were generated
        messages.success(request, f"CSV export vytvořen s {total_sms} SMS zprávami.")
        
        return response

class ClientListView(ManagerGroupRequiredMixin, View):
    """
    View for managers to browse all clients with filtering options.
    
    Allows filtering by region, time period, and viewing detailed analytics
    on client turnover and points across their contract brands.
    """
    template_name = 'manager/client_list.html'
    
    def get(self, request):
        from django.db.models import Sum, Count, F, Q, Value, DecimalField
        from django.db.models.functions import Coalesce
        
        # Get filter parameters
        region_id = request.GET.get('region', '')
        year_from = request.GET.get('year_from', datetime.now().year)
        month_from = request.GET.get('month_from', 1)
        year_to = request.GET.get('year_to', datetime.now().year)
        month_to = request.GET.get('month_to', 12)
        
        try:
            year_from = int(year_from)
            month_from = int(month_from)
            year_to = int(year_to)
            month_to = int(month_to)
        except (ValueError, TypeError):
            # Use default values if conversion fails
            year_from = datetime.now().year
            month_from = 1
            year_to = datetime.now().year
            month_to = 12
        
        # Calculate date range for filtering
        date_from = date(year_from, month_from, 1)
        if month_to == 12:
            date_to = date(year_to + 1, 1, 1) - timedelta(days=1)
        else:
            date_to = date(year_to, month_to + 1, 1) - timedelta(days=1)
        
        # Base query - get all active users that are not staff
        clients = User.objects.filter(is_active=True, is_staff=False)
        
        # Apply region filter if specified
        if region_id and region_id != 'all':
            clients = clients.filter(region_id=region_id)
        
        # Annotate with point data for the period
        clients = clients.annotate(
            confirmed_points=Coalesce(
                Sum('pointstransaction__value', 
                    filter=Q(
                        pointstransaction__status='CONFIRMED',
                        pointstransaction__date__gte=date_from,
                        pointstransaction__date__lte=date_to
                    )),
                Value(0)
            ),
            pending_points=Coalesce(
                Sum('pointstransaction__value', 
                    filter=Q(
                        pointstransaction__status='PENDING',
                        pointstransaction__date__gte=date_from,
                        pointstransaction__date__lte=date_to
                    )),
                Value(0)
            ),
            available_points=Coalesce(
                Sum('pointstransaction__value', 
                    filter=Q(pointstransaction__status='CONFIRMED')),
                Value(0)
            )
        )
        
        # Get all regions for the filter dropdown
        regions = Region.objects.filter(is_active=True).order_by('name')
        
        # For each client, get their contract brands turnover
        client_data = []
        for client in clients:
            # Get active contract for further reference
            try:
                active_contract = UserContract.objects.get(
                    user_id=client,
                    is_active=True
                )
                
                # Get all brands in this contract
                contract_brands = [bb.brand_id for bb in active_contract.brandbonuses.all()]
                
                # Calculate total turnover for the period across contract brands
                total_turnover = InvoiceBrandTurnover.objects.filter(
                    invoice__client_number=client.user_number,
                    invoice__invoice_date__gte=date_from,
                    invoice__invoice_date__lte=date_to,
                    invoice__invoice_type='INVOICE',
                    brand__in=contract_brands
                ).aggregate(
                    total=Coalesce(Sum('amount'), Value(0, output_field=DecimalField()))
                )['total']
                
                # Append to results with the contract and turnover info
                client_data.append({
                    'user': client,
                    'contract': active_contract,
                    'turnover': total_turnover,
                    'brand_count': len(contract_brands)
                })
            except UserContract.DoesNotExist:
                # Client has no active contract
                client_data.append({
                    'user': client,
                    'contract': None,
                    'turnover': 0,
                    'brand_count': 0
                })
        
        # Calculate date ranges for quick filter buttons
        current_year = datetime.now().year
        ytd_from = date(current_year, 1, 1)
        ytd_to = date.today()
        last_year_from = date(current_year - 1, 1, 1)
        last_year_to = date(current_year - 1, 12, 31)
        
        # Prepare context
        context = {
            'clients': client_data,
            'regions': regions,
            'selected_region': region_id,
            'year_from': year_from,
            'month_from': month_from,
            'year_to': year_to,
            'month_to': month_to,
            'date_from': date_from,
            'date_to': date_to,
            'ytd_from': ytd_from,
            'ytd_to': ytd_to,
            'last_year_from': last_year_from,
            'last_year_to': last_year_to,
            'current_year': current_year,
            'months': [(i, date(2000, i, 1).strftime('%B')) for i in range(1, 13)]
        }
        
        return render(request, self.template_name, context)
    
class ClientDetailView(ManagerGroupRequiredMixin, View):
    """
    Detailed view of a client for managers.
    
    Shows complete client information including:
    - Contact details
    - Contract information
    - Turnover by brand
    - Points transactions
    - Reward requests
    """
    template_name = 'manager/client_detail.html'
    
    def get(self, request, pk):
        from django.db.models import Sum, Count, F, Q, Value, DecimalField
        from django.db.models.functions import Coalesce
        import datetime
        
        # Get the client
        client = get_object_or_404(User, pk=pk)
        
        # Get filter parameters for date range
        year_from = request.GET.get('year_from', timezone.now().year)
        month_from = request.GET.get('month_from', 1)
        year_to = request.GET.get('year_to', timezone.now().year)
        month_to = request.GET.get('month_to', 12)
        
        try:
            year_from = int(year_from)
            month_from = int(month_from)
            year_to = int(year_to)
            month_to = int(month_to)
        except (ValueError, TypeError):
            # Use default values if conversion fails
            year_from = timezone.now().year
            month_from = 1
            year_to = timezone.now().year
            month_to = 12
        
        # Calculate date range for filtering
        date_from = date(year_from, month_from, 1)
        if month_to == 12:
            date_to = date(year_to + 1, 1, 1) - timedelta(days=1)
        else:
            date_to = date(year_to, month_to + 1, 1) - timedelta(days=1)
        
        # Get client's active contract
        try:
            active_contract = UserContract.objects.get(
                user_id=client, 
                is_active=True
            )
            contract_brands = [bb.brand_id for bb in active_contract.brandbonuses.all()]
        except UserContract.DoesNotExist:
            active_contract = None
            contract_brands = []
        
        # Get all client's contracts for history
        all_contracts = UserContract.objects.filter(
            user_id=client
        ).order_by('-contract_date_from')
        
        # Get all brands turnover for the selected period
        all_brands = Brand.objects.all()
        brand_turnovers = []
        
        for brand in all_brands:
            # Get invoice turnover for this brand
            invoice_turnover = InvoiceBrandTurnover.objects.filter(
                invoice__client_number=client.user_number,
                invoice__invoice_date__gte=date_from,
                invoice__invoice_date__lte=date_to,
                invoice__invoice_type='INVOICE',
                brand=brand
            ).aggregate(
                total=Coalesce(Sum('amount'), Value(0, output_field=DecimalField()))
            )['total']
            
            # Get credit note turnover for this brand (negative)
            credit_turnover = InvoiceBrandTurnover.objects.filter(
                invoice__client_number=client.user_number,
                invoice__invoice_date__gte=date_from,
                invoice__invoice_date__lte=date_to,
                invoice__invoice_type='CREDIT_NOTE',
                brand=brand
            ).aggregate(
                total=Coalesce(Sum('amount'), Value(0, output_field=DecimalField()))
            )['total']
            
            # Calculate points for this brand in the period
            points = PointsTransaction.objects.filter(
                user=client,
                date__gte=date_from,
                date__lte=date_to,
                brand=brand,
                status='CONFIRMED'
            ).aggregate(
                total=Coalesce(Sum('value'), Value(0))
            )['total']
            
            # Only include brands with some activity
            if invoice_turnover > 0 or credit_turnover > 0 or points != 0:
                # Check if this brand is in the client's contract
                in_contract = brand in contract_brands
                
                brand_turnovers.append({
                    'brand': brand,
                    'invoice_turnover': invoice_turnover,
                    'credit_turnover': credit_turnover,
                    'net_turnover': invoice_turnover - credit_turnover,
                    'points': points,
                    'in_contract': in_contract
                })
        
        # Sort by net turnover
        brand_turnovers.sort(key=lambda x: x['net_turnover'], reverse=True)
        
        # Get point totals
        point_totals = {
            'available': client.get_balance(),
            'period_confirmed': PointsTransaction.objects.filter(
                user=client,
                date__gte=date_from,
                date__lte=date_to,
                status='CONFIRMED'
            ).aggregate(
                total=Coalesce(Sum('value'), Value(0))
            )['total'],
            'period_pending': PointsTransaction.objects.filter(
                user=client,
                date__gte=date_from,
                date__lte=date_to,
                status='PENDING'
            ).aggregate(
                total=Coalesce(Sum('value'), Value(0))
            )['total']
        }
        
        # Get recent transactions
        recent_transactions = PointsTransaction.objects.filter(
            user=client
        ).order_by('-date', '-created_at')[:10]
        
        # Get reward requests
        reward_requests = RewardRequest.objects.filter(
            user=client
        ).order_by('-requested_at')[:10]
        
        # Prepare context
        context = {
            'client': client,
            'active_contract': active_contract,
            'all_contracts': all_contracts,
            'brand_turnovers': brand_turnovers,
            'point_totals': point_totals,
            'recent_transactions': recent_transactions,
            'reward_requests': reward_requests,
            'date_from': date_from,
            'date_to': date_to,
            'year_from': year_from,
            'month_from': month_from,
            'year_to': year_to,
            'month_to': month_to,
            'months': [(i, date(2000, i, 1).strftime('%B')) for i in range(1, 13)]
        }
        
        return render(request, self.template_name, context)

class UserActivityDashboardView(ManagerGroupRequiredMixin, View):
    """
    View for managers to analyze user activity.
    """
    template_name = 'manager/user_activity_dashboard.html'
    
    def get(self, request):
        from django.db.models import Count, Sum
        from django.db.models.functions import TruncDay, TruncMonth
        import datetime
        
        # Get activity for the last 30 days
        thirty_days_ago = timezone.now().date() - timedelta(days=30)
        
        # Daily activity counts
        daily_activity = UserActivity.objects.filter(
            date__gte=thirty_days_ago
        ).annotate(
            day=TruncDay('date')
        ).values('day').annotate(
            users=Count('user', distinct=True),
            visits=Sum('visit_count')
        ).order_by('day')
        
        # Most active users
        most_active_users = UserActivity.objects.filter(
            date__gte=thirty_days_ago
        ).values('user__username', 'user__first_name', 'user__last_name', 'user__email').annotate(
            total_visits=Sum('visit_count')
        ).order_by('-total_visits')[:20]
        
        # Recently active users (last 7 days)
        seven_days_ago = timezone.now().date() - timedelta(days=7)
        recently_active = UserActivity.objects.filter(
            date__gte=seven_days_ago
        ).values('user').distinct().count()
        
        # Total users in system
        total_users = User.objects.count()
        
        # Percent of active users
        active_percent = (recently_active / total_users * 100) if total_users > 0 else 0
        
        context = {
            'daily_activity': daily_activity,
            'most_active_users': most_active_users,
            'recently_active': recently_active,
            'total_users': total_users,
            'active_percent': active_percent,
        }
        
        return render(request, self.template_name, context)
    
class GoalEvaluationView(ManagerGroupRequiredMixin, View):
    """
    Allows managers to evaluate extra goals and award bonus points.
    Similar to transaction approval but for goal achievements.
    """
    template_name = 'manager/goal_evaluation.html'
    
    def _calculate_points_cap(self, goal):
        """
        Calculate the maximum points cap for the contract period.
        Base: 20,000 points for 12 months, proportional for other lengths.
        
        Args:
            goal: UserContractGoal instance
            
        Returns:
            int: Maximum points allowed for the contract period
        """
        total_days = (goal.goal_period_to - goal.goal_period_from).days
        # Approximate months (30.44 days per month on average)
        total_months = total_days / 30.44
        # 20,000 points for 12 months = ~1,667 points per month
        return int(total_months * 1667)

    def _get_total_awarded_points(self, goal, exclude_evaluation=None):
        """
        Get total points already awarded for this goal across all evaluations.
        
        Args:
            goal: UserContractGoal instance
            exclude_evaluation: Optional evaluation to exclude from the sum
            
        Returns:
            int: Total points already awarded
        """
        from django.db.models import Sum
        
        evaluations = goal.evaluations.all()
        if exclude_evaluation:
            evaluations = evaluations.exclude(id=exclude_evaluation.id)
        
        return evaluations.aggregate(
            total=Sum('bonus_points')
        )['total'] or 0

    def _apply_points_cap(self, points, goal, existing_points=0):
        """
        Apply the points cap, considering already awarded points.
        
        Args:
            points: Points to be awarded
            goal: UserContractGoal instance
            existing_points: Points already awarded
            
        Returns:
            int: The capped points amount that can be awarded
        """
        cap = self._calculate_points_cap(goal)
        total_with_new = existing_points + points
        
        if total_with_new > cap:
            # Can only award up to the cap
            return max(0, cap - existing_points)
        return points

    def _determine_evaluation_result(self, goal, start_date, end_date, actual_turnover, targets, is_final):
        """
        Determine evaluation type, bonus points, and achievement status.
        Implements the business logic for different evaluation scenarios with point caps.
        
        Business Rules:
        1. Milestone evaluations: Award points if period target is met
        2. Final evaluations with recovery: 
        - First evaluate the final period normally
        - Then check if full year target is met
        - If yes, calculate total year points and subtract already awarded
        3. All points are subject to the proportional annual cap
        
        Args:
            goal: UserContractGoal instance
            start_date: Period start date
            end_date: Period end date
            actual_turnover: Actual turnover achieved
            targets: Dict with 'goal_value' and 'goal_base' for the period
            is_final: Whether this is the final evaluation period
            
        Returns:
            tuple: (evaluation_type, bonus_points, is_achieved)
        """
        # Get total points already awarded for this goal
        already_awarded = self._get_total_awarded_points(goal)
        
        # Check if this period's target was achieved
        period_achieved = actual_turnover >= targets['goal_value']
        
        if not is_final:
            # Standard milestone evaluation
            if period_achieved:
                # Calculate points: half of the increase from base to goal
                raw_points = int((targets['goal_value'] - targets['goal_base']) * goal.bonus_percentage)
                # Apply cap considering already awarded points
                bonus_points = self._apply_points_cap(raw_points, goal, already_awarded)
                return 'MILESTONE', max(0, bonus_points), True
            else:
                return 'MILESTONE', 0, False
        
        # For final evaluation, check recovery scenarios
        if goal.allow_full_period_recovery:
            # First, evaluate the final period as a milestone
            period_points = 0
            if period_achieved:
                raw_points = int((float(actual_turnover) - targets['goal_base']) * goal.bonus_percentage)
                period_points = self._apply_points_cap(raw_points, goal, already_awarded)
            
            # Get all previous evaluations to calculate total milestone points
            previous_evals = goal.evaluations.filter(
                period_end__lt=end_date
            ).order_by('period_end')
            
            # Calculate total points from all previous milestones
            previous_milestone_points = previous_evals.aggregate(
                total=Sum('bonus_points')
            )['total'] or 0
            
            # Check if we should attempt recovery
            # Get full period actual turnover
            full_actual = calculate_turnover_for_goal(
                goal.user_contract.user_id,
                goal.brands.all(),
                goal.goal_period_from,
                goal.goal_period_to
            )
            
            if full_actual >= goal.goal_value:
                # Recovery scenario - calculate total points for the year
                # Points = half of (actual turnover - base)
                total_year_points = int(float((full_actual - goal.goal_base)) * goal.bonus_percentage)
                
                # Apply the annual cap
                capped_year_points = min(total_year_points, self._calculate_points_cap(goal))
                
                # Calculate how many points we need to add (recovery points)
                # This is the difference between what they should get for the year 
                # and what they already got from milestones
                recovery_points = capped_year_points - previous_milestone_points
                
                if recovery_points > period_points:
                    # Recovery gives more points than just the final period
                    return 'RECOVERY', max(0, recovery_points), True
                else:
                    # Final period evaluation gives more or equal points
                    return 'FINAL', max(0, period_points), period_achieved
            else:
                # No recovery possible, just evaluate the final period
                return 'FINAL', max(0, period_points), period_achieved
        else:
            # Simple evaluation without recovery option
            if period_achieved:
                raw_points = int((float(actual_turnover) - targets['goal_base']) * goal.bonus_percentage)
                bonus_points = self._apply_points_cap(raw_points, goal, already_awarded)
                return 'FINAL', max(0, bonus_points), True
            else:
                return 'FINAL', 0, False

    def get(self, request):
        """
        Display pending evaluations or handle export request.
        Added: ?export=preview parameter to download Excel preview.
        """
        from django.db.models import Q, Exists, OuterRef
        today = timezone.now().date()
        
        # Get filter parameters
        evaluation_type = request.GET.get('type', 'pending')
        region_id = request.GET.get('region', '')
        
        # Base query - get goals with evaluation periods that have ended
        goals_query = UserContractGoal.objects.filter(
            goal_period_from__lte=today
        ).select_related('user_contract__user_id')
        
        # Apply region filter if specified
        if region_id and region_id != 'all':
            goals_query = goals_query.filter(
                user_contract__user_id__region_id=region_id
            )
        
        # Process each goal to find pending evaluations
        pending_evaluations = []
        
        for goal in goals_query:
            periods = goal.get_evaluation_periods()
            
            for start_date, end_date, is_final in periods:
                # Skip future periods
                if end_date > today:
                    continue
                
                # Check if already evaluated
                existing_evaluation = goal.evaluations.filter(
                    period_start=start_date,
                    period_end=end_date
                ).first()
                
                if evaluation_type == 'pending' and existing_evaluation:
                    continue
                elif evaluation_type == 'evaluated' and not existing_evaluation:
                    continue
                
                # Calculate turnover and targets
                targets = goal.get_period_targets(start_date, end_date)
                actual_turnover = calculate_turnover_for_goal(
                    goal.user_contract.user_id,
                    goal.brands.all(),
                    start_date,
                    end_date
                )
                
                # Calculate full year turnover for diagnostics
                full_year_actual = calculate_turnover_for_goal(
                    goal.user_contract.user_id,
                    goal.brands.all(),
                    goal.goal_period_from,
                    goal.goal_period_to
                )
                
                # Get already awarded points
                already_awarded = self._get_total_awarded_points(goal)
                
                # Calculate points cap
                points_cap = self._calculate_points_cap(goal)
                
                # Determine evaluation type and potential bonus
                eval_type, bonus_points, is_achieved = self._determine_evaluation_result(
                    goal, start_date, end_date, actual_turnover, targets, is_final
                )
                
                pending_evaluations.append({
                    'goal': goal,
                    'user': goal.user_contract.user_id,
                    'period_start': start_date,
                    'period_end': end_date,
                    'is_final': is_final,
                    'actual_turnover': actual_turnover,
                    'target_turnover': targets['goal_value'],
                    'baseline_turnover': targets['goal_base'],
                    'evaluation_type': eval_type,
                    'bonus_points': bonus_points,
                    'is_achieved': is_achieved,
                    'existing_evaluation': existing_evaluation,
                    'brands': list(goal.brands.all()),
                    # Additional diagnostic fields for export
                    'full_year_actual': full_year_actual,
                    'full_year_goal': goal.goal_value,
                    'full_year_base': goal.goal_base,
                    'full_year_met': full_year_actual >= goal.goal_value,
                    'points_cap': points_cap,
                    'already_awarded': already_awarded,
                    'goal_period_from': goal.goal_period_from,
                    'goal_period_to': goal.goal_period_to,
                })
        
        # Check if this is an export request
        if request.GET.get('export') == 'preview':
            return self.generate_evaluation_export(pending_evaluations)
        
        # Get regions for filter
        regions = Region.objects.filter(is_active=True).order_by('name')
        
        # Sort evaluations by user and date
        pending_evaluations.sort(key=lambda x: (x['user'].last_name, x['period_end']))
        
        context = {
            'evaluations': pending_evaluations,
            'regions': regions,
            'selected_region': region_id,
            'evaluation_type': evaluation_type,
            'today': today
        }
        
        return render(request, self.template_name, context)

    def generate_evaluation_export(self, evaluations):
        """
        Generate an Excel export of pending evaluations for review.
        Includes diagnostic columns to help verify the calculation logic.
        
        Args:
            evaluations: List of evaluation dictionaries from get()
            
        Returns:
            HttpResponse: Excel file download
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Evaluation Preview'
        
        # Define headers - including diagnostic columns
        headers = [
            'Client ID',
            'Client Name',
            'Region',
            'Brands',
            'Period Start',
            'Period End',
            'Is Final Period',
            'Period Target',
            'Period Base',
            'Period Actual',
            'Period Achieved',
            'Full Year Goal',
            'Full Year Base', 
            'Full Year Actual',
            'Full Year Met',
            'Evaluation Type',
            'Points to Award',
            'Already Awarded',
            'Points Cap',
            'Existing Evaluation',
        ]
        
        # Style for headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Highlight colors for key columns
        warning_fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')  # Light red
        success_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Write data rows
        for row_idx, eval_data in enumerate(evaluations, 2):
            user = eval_data['user']
            brand_names = ', '.join([b.name for b in eval_data['brands']])
            
            values = [
                user.user_number,
                f"{user.first_name} {user.last_name}",
                user.region.name if user.region else 'No Region',
                brand_names,
                eval_data['period_start'],
                eval_data['period_end'],
                'Yes' if eval_data['is_final'] else 'No',
                float(eval_data['target_turnover']),
                float(eval_data['baseline_turnover']),
                float(eval_data['actual_turnover']),
                'Yes' if eval_data['is_achieved'] else 'No',
                float(eval_data['full_year_goal']),
                float(eval_data['full_year_base']),
                float(eval_data['full_year_actual']),
                'Yes' if eval_data['full_year_met'] else 'No',
                eval_data['evaluation_type'],
                eval_data['bonus_points'],
                eval_data['already_awarded'],
                eval_data['points_cap'],
                'Yes' if eval_data['existing_evaluation'] else 'No',
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                
                # Highlight rows where points are being awarded but full year not met
                # This is the scenario you're concerned about
                if col == 17 and eval_data['bonus_points'] > 0 and not eval_data['full_year_met']:
                    cell.fill = warning_fill
                
                # Highlight full year met column
                if col == 15:
                    if eval_data['full_year_met']:
                        cell.fill = success_fill
                    else:
                        cell.fill = warning_fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create HTTP response
        filename = f'goal_evaluation_preview_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @transaction.atomic
    def post(self, request):
        """Process selected goal evaluations and create bonus transactions."""
        evaluations_to_process = request.POST.getlist('evaluate')
        
        success_count = 0
        total_points = 0
        
        for eval_key in evaluations_to_process:
            # Parse the evaluation key (format: "goal_id:start_date:end_date")
            try:
                goal_id, start_str, end_str = eval_key.split(':')
                goal = UserContractGoal.objects.get(id=goal_id)
                start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
            except (ValueError, UserContractGoal.DoesNotExist):
                continue
            
            # Check if already evaluated
            if goal.evaluations.filter(period_start=start_date, period_end=end_date).exists():
                continue
            
            # Calculate evaluation details
            targets = goal.get_period_targets(start_date, end_date)
            actual_turnover = calculate_turnover_for_goal(
                goal.user_contract.user_id,
                goal.brands.all(),
                start_date,
                end_date
            )
            
            eval_type, bonus_points, is_achieved = self._determine_evaluation_result(
                goal, start_date, end_date, actual_turnover, targets, 
                end_date == goal.goal_period_to
            )
            
            # Create evaluation record
            evaluation = GoalEvaluation.objects.create(
                goal=goal,
                evaluation_date=timezone.now().date(),
                period_start=start_date,
                period_end=end_date,
                actual_turnover=actual_turnover,
                target_turnover=targets['goal_value'],
                baseline_turnover=targets['goal_base'],
                is_achieved=is_achieved,
                bonus_points=bonus_points,
                evaluation_type=eval_type,
                evaluated_by=request.user
            )
            
            # Create points transaction if bonus points awarded
            if bonus_points > 0:
                transaction = PointsTransaction.objects.create(
                    user=goal.user_contract.user_id,
                    value=bonus_points,
                    date=end_date,
                    description=f"Extra bonus za období od {start_date} do {end_date}",
                    type='EXTRA_POINTS',
                    status='CONFIRMED'
                )
                evaluation.points_transaction = transaction
                evaluation.save()
                
                total_points += bonus_points
            
            success_count += 1
        
        messages.success(
            request,
            f"Successfully evaluated {success_count} goal periods, "
            f"awarding {total_points} bonus points in total."
        )
        
        return redirect('goal_evaluation')
    
class GoalsOverviewView(ManagerGroupRequiredMixin, ListView):
    """
    Manager view showing overview of all extra goals with filtering and progress tracking.
    Now includes export functionality for full contract data and current period data.
    """
    template_name = 'manager/goals_overview.html'
    context_object_name = 'goal_data'
    paginate_by = 300
    
    def get(self, request, *args, **kwargs):
        # Check if this is an export request
        export_type = request.GET.get('export')
        if export_type in ['full', 'current']:
            return self.handle_export(export_type)
        
        # Otherwise, handle normal page request
        return super().get(request, *args, **kwargs)
    
    def handle_export(self, export_type):
        """
        Handle export requests for either full contract data or current period data.
        
        Args:
            export_type (str): Either 'full' or 'current'
            
        Returns:
            HttpResponse: Excel file download response
        """
        today = timezone.now().date()
        
        # Get the same queryset as the main view but without pagination
        queryset = self.get_base_queryset()
        
        # Process the data for export
        export_data = []
        for goal in queryset:
            user = goal.user_contract.user_id
            
            if export_type == 'full':
                # Full export: entire contract period
                export_data.append(self.get_full_export_row(goal, user, today))
            else:
                # Current period export: current milestone only
                current_period_data = self.get_current_period_data(goal, user, today)
                if current_period_data:
                    export_data.append(current_period_data)
        
        # Generate Excel file
        return self.generate_excel_response(export_data, export_type)
    
    def get_base_queryset(self):
        """
        Get the base queryset with all filters applied, similar to get_queryset 
        but without pagination limits.
        """
        today = timezone.now().date()
        
        queryset = UserContractGoal.objects.filter(
            goal_period_from__lte=today,
            goal_period_to__gte=today
        ).select_related(
            'user_contract__user_id__region'
        ).prefetch_related(
            'brands',
            'user_contract__brandbonuses__brand_id'
        )
        
        # Apply region filter
        region_id = self.request.GET.get('region')
        if region_id and region_id != 'all':
            queryset = queryset.filter(
                user_contract__user_id__region_id=region_id
            )
        
        # Apply brand filter
        brand_id = self.request.GET.get('brand')
        if brand_id and brand_id != 'all':
            queryset = queryset.filter(
                brands__id=brand_id
            ).distinct()
        
        return queryset.order_by('user_contract__user_id__last_name')
    
    def get_full_export_row(self, goal, user, today):
        """
        Generate a row of data for the full export (entire contract period).
        """
        # Calculate total turnover for the entire goal period
        current_turnover = calculate_turnover_for_goal(
            user,
            goal.brands.all(),
            goal.goal_period_from,
            min(today, goal.goal_period_to)
        )
        
        # Calculate percentage and remaining turnover
        percentage = (float(current_turnover) / goal.goal_value) if goal.goal_value > 0 else 0
        remaining_turnover = max(0, goal.goal_value - float(current_turnover))

        # Get brand names as comma-separated string
        brand_names = ', '.join([brand.name for brand in goal.brands.all()])
        
        return {
            'client_id': user.user_number,
            'client_name': f"{user.first_name} {user.last_name}",
            'client_region': user.region.name if user.region else 'No Region',
            'goal_brands': brand_names,
            'goal_period_from': goal.goal_period_from,
            'goal_period_to': goal.goal_period_to,
            'goal_value': goal.goal_value,
            'current_turnover': float(current_turnover),
            'percentage_of_goal': round(percentage, 2),
            'turnover_remaining': float(remaining_turnover)
        }
    
    def get_current_period_data(self, goal, user, today):
        """
        Generate a row of data for the current period export (current milestone).
        """
        # Find the current evaluation period
        periods = goal.get_evaluation_periods()
        current_period = None
        
        for start, end, is_final in periods:
            if start <= today <= end:
                current_period = (start, end, is_final)
                break
        
        # If no current period found, skip this goal
        if not current_period:
            return None
        
        start_date, end_date, is_final = current_period
        
        # Calculate targets for this period
        targets = goal.get_period_targets(start_date, end_date)
        
        # Calculate actual turnover for this period
        period_turnover = calculate_turnover_for_goal(
            user,
            goal.brands.all(),
            start_date,
            min(today, end_date)
        )
        
        # Calculate percentage and remaining turnover
        percentage = (float(period_turnover) / targets['goal_value']) if targets['goal_value'] > 0 else 0
        remaining_turnover = max(0, targets['goal_value'] - float(period_turnover))
        
        # Get brand names as comma-separated string
        brand_names = ', '.join([brand.name for brand in goal.brands.all()])
        
        return {
            'client_id': user.user_number,
            'client_name': f"{user.first_name} {user.last_name}",
            'client_region': user.region.name if user.region else 'No Region',
            'goal_brands': brand_names,
            'milestone_period_from': start_date,
            'milestone_period_to': end_date,
            'period_goal_value': targets['goal_value'],
            'period_turnover': float(period_turnover),
            'percentage_of_goal': round(percentage, 2),
            'turnover_remaining': float(remaining_turnover)
        }
    
    def generate_excel_response(self, data, export_type):
        """
        Generate an Excel file from the export data and return as HTTP response.
        """
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        
        if export_type == 'full':
            ws.title = 'Full Goals Export'
            headers = [
                'Client ID', 'Client Name', 'Client Region', 'Goal Brands',
                'Goal Period From', 'Goal Period To', 'Goal Value', 'Current Turnover', 
                'Percentage of Goal', 'Turnover Remaining'
            ]
            filename = f'goals_full_export_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
        else:
            ws.title = 'Current Period Export'
            headers = [
                'Client ID', 'Client Name', 'Client Region', 'Goal Brands',
                'Milestone Period From', 'Milestone Period To', 'Period Goal Value',
                'Period Turnover', 'Percentage of Goal', 'Turnover Remaining'
            ]
            filename = f'goals_current_period_export_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Add data rows
        for row_idx, row_data in enumerate(data, 2):
            if export_type == 'full':
                values = [
                    row_data['client_id'],
                    row_data['client_name'],
                    row_data['client_region'],
                    row_data['goal_brands'],
                    row_data['goal_period_from'],
                    row_data['goal_period_to'],
                    row_data['goal_value'],
                    row_data['current_turnover'],
                    row_data['percentage_of_goal'],
                    row_data['turnover_remaining']
                ]
            else:
                values = [
                    row_data['client_id'],
                    row_data['client_name'],
                    row_data['client_region'],
                    row_data['goal_brands'],
                    row_data['milestone_period_from'],
                    row_data['milestone_period_to'],
                    row_data['period_goal_value'],
                    row_data['period_turnover'],
                    row_data['percentage_of_goal'],
                    row_data['turnover_remaining']
                ]
            
            for col, value in enumerate(values, 1):
                ws.cell(row=row_idx, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create HTTP response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response

    # Keep the existing get_queryset and get_context_data methods unchanged
    def get_queryset(self):
        # ... existing code remains the same
        return self.get_base_queryset()
    
    def get_context_data(self, **kwargs):
        # ... existing code remains the same
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()
        
        # Process each goal to add calculated fields
        processed_goals = []
        for goal in context['goal_data']:
            user = goal.user_contract.user_id
            
            # Calculate current turnover
            current_turnover = calculate_turnover_for_goal(
                user,
                goal.brands.all(),
                goal.goal_period_from,
                min(today, goal.goal_period_to)  # Don't count future dates
            )
            
            # Calculate ideal turnover (linear progression)
            total_days = (goal.goal_period_to - goal.goal_period_from).days + 1
            elapsed_days = (today - goal.goal_period_from).days + 1
            
            # Ensure we don't exceed 100% for ideal calculation
            if today >= goal.goal_period_to:
                progress_ratio = 1.0
                elapsed_days = total_days
            else:
                progress_ratio = elapsed_days / total_days
            
            ideal_turnover = goal.goal_value * progress_ratio
            
            # Calculate percentage of goal achieved
            goal_percentage = (float(current_turnover) / goal.goal_value * 100) if goal.goal_value > 0 else 0
            
            # Determine if on track (within 5% of ideal)
            ideal_percentage = (float(current_turnover) / ideal_turnover * 100) if ideal_turnover > 0 else 0
            if ideal_percentage >= 95:
                track_status = 'on_track'
            elif ideal_percentage >= 85:
                track_status = 'slightly_behind'
            else:
                track_status = 'behind'
            
            processed_goals.append({
                'goal': goal,
                'user': user,
                'contract': goal.user_contract,
                'brands': list(goal.brands.all()),
                'current_turnover': current_turnover,
                'ideal_turnover': ideal_turnover,
                'goal_percentage': goal_percentage,
                'ideal_percentage': ideal_percentage,
                'track_status': track_status,
                'days_elapsed': elapsed_days,
                'days_total': total_days,
                'progress_ratio': progress_ratio * 100  # As percentage
            })
        
        # Replace the paginated data with our processed data
        context['goal_data'] = processed_goals
        
        # Add filter options
        context['regions'] = Region.objects.filter(is_active=True).order_by('name')
        context['brands'] = Brand.objects.all().order_by('name')
        
        # Add selected filters
        context['selected_region'] = self.request.GET.get('region', 'all')
        context['selected_brand'] = self.request.GET.get('brand', 'all')
        
        # Summary statistics
        if processed_goals:
            context['summary'] = {
                'total_goals': len(processed_goals),
                'on_track': sum(1 for g in processed_goals if g['track_status'] == 'on_track'),
                'slightly_behind': sum(1 for g in processed_goals if g['track_status'] == 'slightly_behind'),
                'behind': sum(1 for g in processed_goals if g['track_status'] == 'behind'),
                'avg_percentage': sum(g['goal_percentage'] for g in processed_goals) / len(processed_goals)
            }
        else:
            context['summary'] = {
                'total_goals': 0,
                'on_track': 0,
                'slightly_behind': 0,
                'behind': 0,
                'avg_percentage': 0
            }
        
        return context
    

class EnhancedRewardRequestListView(ManagerGroupRequiredMixin, View):
    """
    Enhanced reward request management interface with comprehensive features.
    Combines list view, analytics, and bulk operations in a single interface.
    """
    template_name = 'manager/reward_requests_enhanced.html'
    
    def get(self, request):
        # Get filter parameters
        status_filter = request.GET.get('status', '')
        search_query = request.GET.get('search', '')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        client_filter = request.GET.get('client', '')
        sort_by = request.GET.get('sort', '-requested_at')
        
        # Build base queryset with optimized joins
        queryset = RewardRequest.objects.select_related('user', 'user__region').prefetch_related(
            'rewardrequestitem_set__reward',
            'pointstransaction_set'
        )
        
        # Apply filters
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        if search_query:
            queryset = queryset.filter(
                Q(user__first_name__icontains=search_query) |
                Q(user__last_name__icontains=search_query) |
                Q(user__user_number__icontains=search_query) |
                Q(id__icontains=search_query)
            )
        
        if date_from:
            queryset = queryset.filter(requested_at__gte=date_from)
        
        if date_to:
            # Add 1 day to include the entire end date
            end_date = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
            queryset = queryset.filter(requested_at__lt=end_date)
        
        if client_filter:
            queryset = queryset.filter(user_id=client_filter)
        
        # Apply sorting
        queryset = queryset.order_by(sort_by)
        
        # Get all requests for the current filters (before pagination)
        all_requests = list(queryset)
        
        # Paginate for display (but keep all_requests for analytics)
        paginator = Paginator(queryset, 50)  # Show 50 per page
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        
        # Calculate analytics for filtered results
        analytics = self._calculate_analytics(all_requests)
        
        # Get status counts for tabs
        status_counts = self._get_status_counts()
        
        # Get list of clients for filter dropdown
        clients_with_requests = User.objects.filter(
            rewardrequest__isnull=False
        ).distinct().order_by('first_name', 'last_name')
        
        # Prepare reward inventory summary
        reward_inventory = self._get_reward_inventory(status_filter or 'PENDING')
        
        context = {
            'requests': page_obj,
            'all_requests': all_requests,  # For expandable details
            'page_obj': page_obj,
            'status_filter': status_filter,
            'search_query': search_query,
            'date_from': date_from,
            'date_to': date_to,
            'client_filter': client_filter,
            'sort_by': sort_by,
            'status_counts': status_counts,
            'analytics': analytics,
            'reward_inventory': reward_inventory,
            'clients_with_requests': clients_with_requests,
            'request_statuses': RewardRequest.REQUEST_STATUS,
        }
        
        return render(request, self.template_name, context)
    
    def post(self, request):
        """Handle bulk operations"""
        action = request.POST.get('action')
        selected_ids = request.POST.getlist('selected_requests')
        
        if not selected_ids:
            messages.warning(request, "No requests selected for bulk operation.")
            return redirect('enhanced_reward_requests')
        
        if action == 'bulk_status_update':
            new_status = request.POST.get('new_status')
            if new_status:
                self._bulk_update_status(selected_ids, new_status)
                messages.success(request, f"Updated {len(selected_ids)} requests to {new_status}")
        
        elif action == 'bulk_export':
            # Generate export file
            return self._generate_bulk_export(selected_ids)
        
        return redirect('enhanced_reward_requests')
    
    def _calculate_analytics(self, requests):
        """Calculate comprehensive analytics for the filtered requests"""
        if not requests:
            return {
                'total_requests': 0,
                'total_points': 0,
                'avg_points': 0,
                'status_breakdown': {},
                'top_rewards': [],
                'recent_activity': []
            }
        
        total_points = sum(r.total_points for r in requests)
        
        # Status breakdown
        status_breakdown = {}
        for status_code, status_label in RewardRequest.REQUEST_STATUS:
            count = sum(1 for r in requests if r.status == status_code)
            if count > 0:
                status_breakdown[status_label] = {
                    'count': count,
                    'points': sum(r.total_points for r in requests if r.status == status_code)
                }
        
        # Top requested rewards
        reward_counts = {}
        for request in requests:
            for item in request.rewardrequestitem_set.all():
                reward_key = (item.reward.id, item.reward.name, item.reward.abra_code)
                if reward_key not in reward_counts:
                    reward_counts[reward_key] = {'quantity': 0, 'points': 0, 'requests': 0}
                reward_counts[reward_key]['quantity'] += item.quantity
                reward_counts[reward_key]['points'] += item.quantity * item.point_cost
                reward_counts[reward_key]['requests'] += 1
        
        top_rewards = sorted(
            [{'id': k[0], 'name': k[1], 'code': k[2], **v} for k, v in reward_counts.items()],
            key=lambda x: x['quantity'],
            reverse=True
        )[:10]
        
        # Recent activity (last 7 days)
        seven_days_ago = timezone.now() - timedelta(days=7)
        recent_requests = [r for r in requests if r.requested_at >= seven_days_ago]
        
        return {
            'total_requests': len(requests),
            'total_points': total_points,
            'avg_points': total_points // len(requests) if requests else 0,
            'status_breakdown': status_breakdown,
            'top_rewards': top_rewards,
            'recent_count': len(recent_requests),
            'recent_points': sum(r.total_points for r in recent_requests)
        }
    
    def _get_status_counts(self):
        """Get counts for each status for the tab navigation"""
        counts = RewardRequest.objects.values('status').annotate(count=Count('id'))
        status_dict = {item['status']: item['count'] for item in counts}
        
        # Include total
        total = sum(status_dict.values())
        
        result = [('', 'All', total)]
        for status_code, status_label in RewardRequest.REQUEST_STATUS:
            count = status_dict.get(status_code, 0)
            result.append((status_code, status_label, count))
        
        return result
    
    def _get_reward_inventory(self, status='PENDING'):
        """Get reward inventory summary for a specific status"""
        items = RewardRequestItem.objects.filter(
            reward_request__status=status
        ).values(
            'reward__id',
            'reward__name',
            'reward__abra_code',
            'reward__point_cost'
        ).annotate(
            total_quantity=Sum('quantity'),
            request_count=Count('reward_request', distinct=True),
            total_points=Sum(F('quantity') * F('point_cost'))
        ).order_by('-total_quantity')[:200]
        
        return items
    
    @transaction.atomic
    def _bulk_update_status(self, request_ids, new_status):
        """Update status for multiple requests"""
        requests = RewardRequest.objects.filter(id__in=request_ids)
        
        for reward_request in requests:
            old_status = reward_request.status
            reward_request.status = new_status
            reward_request.save()
            
            # Update associated transactions
            self._update_point_transaction(reward_request, old_status, new_status)
    
    def _update_point_transaction(self, reward_request, old_status, new_status):
        """Update the point transaction to match the current state of the request"""
        try:
            transaction = PointsTransaction.objects.get(
                reward_request=reward_request,
                type='REWARD_CLAIM'
            )
        except PointsTransaction.DoesNotExist:
            logger.warning(f"No transaction found for reward request {reward_request.id}")
            return
        except PointsTransaction.MultipleObjectsReturned:
            logger.error(f"Multiple transactions found for reward request {reward_request.id}")
            return
        
        # Update transaction status based on request status
        if new_status in ['REJECTED', 'CANCELLED']:
            transaction.status = 'CANCELLED'
        elif old_status in ['REJECTED', 'CANCELLED'] and new_status not in ['REJECTED', 'CANCELLED']:
            transaction.status = 'CONFIRMED'
        
        # Ensure transaction amount matches request total
        if transaction.status == 'CONFIRMED':
            transaction.value = -reward_request.total_points
        
        transaction.save()
    
    def _generate_bulk_export(self, request_ids):
        """Generate Excel export for selected requests"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reward Requests"
        
        # Headers
        headers = ['Request ID', 'Client Number', 'Client Name', 'Status', 'Total Points', 
                  'Requested Date', 'Items', 'Note']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="6B9AC4", end_color="6B9AC4", fill_type="solid")
        
        # Data
        requests = RewardRequest.objects.filter(id__in=request_ids).select_related('user')
        for row, request in enumerate(requests, 2):
            items_str = ', '.join([
                f"{item.quantity}x {item.reward.name}"
                for item in request.rewardrequestitem_set.all()
            ])
            
            ws.cell(row=row, column=1, value=request.id)
            ws.cell(row=row, column=2, value=request.user.user_number)
            ws.cell(row=row, column=3, value=f"{request.user.first_name} {request.user.last_name}")
            ws.cell(row=row, column=4, value=request.get_status_display())
            ws.cell(row=row, column=5, value=request.total_points)
            ws.cell(row=row, column=6, value=request.requested_at.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row, column=7, value=items_str)
            ws.cell(row=row, column=8, value=request.note or '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=reward_requests_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
        
        wb.save(response)
        return response


class RewardRequestQuickEditView(ManagerGroupRequiredMixin, View):
    """Handle quick inline editing of reward requests"""
    
    @transaction.atomic
    def post(self, request, pk):
        reward_request = get_object_or_404(RewardRequest, pk=pk)
        
        # Quick status update
        new_status = request.POST.get('status')
        if new_status and new_status in [s[0] for s in RewardRequest.REQUEST_STATUS]:
            old_status = reward_request.status
            reward_request.status = new_status
            
            # Update manager message if provided
            manager_message = request.POST.get('manager_message', '')
            if manager_message:
                reward_request.description = manager_message
            
            reward_request.save()
            
            # Update associated transaction
            self._update_point_transaction(reward_request, old_status, new_status)
            
            messages.success(request, f"Request #{pk} updated to {reward_request.get_status_display()}")
        
        # Rebuild the query string from the filter_ prefixed parameters
        query_params = {}
        for key, value in request.POST.items():
            if key.startswith('filter_'):
                actual_key = key.replace('filter_', '')
                query_params[actual_key] = value
        
        # Build the redirect URL
        base_url = reverse('enhanced_reward_requests')
        if query_params:
            from urllib.parse import urlencode
            redirect_url = f"{base_url}?{urlencode(query_params)}"
        else:
            redirect_url = base_url
        
        return HttpResponseRedirect(redirect_url)
    
    def _update_point_transaction(self, reward_request, old_status, new_status):
        """Update the point transaction to match the current state of the request"""
        try:
            transaction = PointsTransaction.objects.get(
                reward_request=reward_request,
                type='REWARD_CLAIM'
            )
        except PointsTransaction.DoesNotExist:
            logger.warning(f"No transaction found for reward request {reward_request.id}")
            return
        except PointsTransaction.MultipleObjectsReturned:
            logger.error(f"Multiple transactions found for reward request {reward_request.id}")
            return
        
        # Update transaction status based on request status
        if new_status in ['REJECTED', 'CANCELLED']:
            transaction.status = 'CANCELLED'
        elif old_status in ['REJECTED', 'CANCELLED'] and new_status not in ['REJECTED', 'CANCELLED']:
            transaction.status = 'CONFIRMED'
        
        # Ensure transaction amount matches request total
        if transaction.status == 'CONFIRMED':
            transaction.value = -reward_request.total_points
        
        transaction.save()

class ClientCreateView(ManagerGroupRequiredMixin, View):
    """
    View for creating a new client with contract, optional goal, and
    optional retroactive transaction processing.
    """
    
    template_name = 'manager/client_create.html'
    
    def get(self, request):
        """Display the empty client creation form."""
        form = ClientCreationForm()
        return render(request, self.template_name, {'form': form})
    
    def post(self, request):
        """Process the client creation form submission."""
        form = ClientCreationForm(request.POST)
        
        if form.is_valid():
            try:
                # The form's save method handles the entire creation process
                # including optional retroactive transaction processing
                user, transaction_stats = form.save()
                
                # Build success message
                success_message = (
                    f'Client "{user.get_full_name() or user.username}" has been created successfully. '
                    f'Default password is set to their customer number: {user.user_number}'
                )
                
                # Add transaction processing results to message if applicable
                if transaction_stats:
                    success_message += self._format_transaction_stats(transaction_stats)
                
                messages.success(request, success_message)
                
                # Log any errors from transaction processing
                if transaction_stats and transaction_stats.get('errors'):
                    for error in transaction_stats['errors']:
                        logger.error(f"Transaction processing error: {error}")
                        messages.warning(request, f"Warning: {error}")
                
                # Redirect to the client detail page
                return redirect('manager_client_detail', pk=user.id)
                
            except Exception as e:
                logger.error(f"Error creating client: {str(e)}", exc_info=True)
                messages.error(
                    request,
                    f'Error creating client: {str(e)}'
                )
        else:
            # Form validation failed
            messages.error(
                request,
                'Please correct the errors below.'
            )
        
        # Re-display form with errors
        return render(request, self.template_name, {'form': form})
    
    def _format_transaction_stats(self, stats):
        """
        Format transaction processing statistics into a readable message.
        
        Args:
            stats (dict): Statistics from retroactive transaction processing
            
        Returns:
            str: Formatted message describing the results
        """
        message_parts = []
        
        if stats['invoices_found'] > 0:
            message_parts.append(
                f"\n\nHistorical Transaction Processing:"
                f"\n- Found {stats['invoices_found']} historical invoice(s)"
            )
            
            if stats['invoices_processed'] > 0:
                message_parts.append(
                    f"- Processed {stats['invoices_processed']} invoice(s)"
                )
            
            if stats['transactions_created'] > 0:
                message_parts.append(
                    f"- Created {stats['transactions_created']} new transaction(s)"
                )
            
            if stats['transactions_skipped'] > 0:
                message_parts.append(
                    f"- Skipped {stats['transactions_skipped']} duplicate transaction(s)"
                )
            
            if stats['brands_without_bonus'] > 0:
                message_parts.append(
                    f"- Skipped {stats['brands_without_bonus']} brand(s) not in contract"
                )
        else:
            message_parts.append(
                "\n\nNo historical invoices found for this client number."
            )
        
        return ''.join(message_parts)
