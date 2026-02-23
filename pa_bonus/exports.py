import io
from django.shortcuts import get_object_or_404
from django.db import transaction
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pa_bonus.models import RewardRequest

def generate_telemarketing_export(request_id):
    """
    Generate an Excel file for telemarketing from a specific reward request.
    
    Args:
        request_id (int): ID of the RewardRequest to export
        
    Returns:
        bytes: Excel file content or None if the request is not found or not in ACCEPTED status
    """

    def apply_item_styling(row_num):
        # Data cells style (light green background for columns B, E, F)
        data_font = Font(name='Arial', size=11)
        data_fill = PatternFill(start_color='E6FFE6', end_color='E6FFE6', fill_type='solid')
        data_alignment = Alignment(horizontal='left', vertical='center')
        data_alignment_center = Alignment(horizontal='center', vertical='center')

        ws[f'B{row_num}'].fill = data_fill
        ws[f'B{row_num}'].alignment = data_alignment
        ws[f'B{row_num}'].font = data_font
        
        ws[f'C{row_num}'].alignment = data_alignment
        ws[f'C{row_num}'].font = data_font

        ws[f'D{row_num}'].fill = data_fill
        ws[f'D{row_num}'].alignment = data_alignment_center
        ws[f'D{row_num}'].font = data_font
        
        ws[f'E{row_num}'].fill = data_fill
        ws[f'E{row_num}'].alignment = data_alignment_center
        ws[f'E{row_num}'].font = data_font
        
        ws[f'F{row_num}'].fill = data_fill
        ws[f'F{row_num}'].alignment = data_alignment_center
        ws[f'F{row_num}'].font = data_font

    try:
        # Get the reward request
        request = get_object_or_404(RewardRequest, pk=request_id)
        
        # Check if it's in ACCEPTED status
        if request.status != 'ACCEPTED':
            return None
            
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Request {request_id}"
        
        # Define styles
        # First row style (red, bold, size 18)
        first_row_font = Font(name='Arial', size=18, bold=True, color='FF0000')
        first_row_alignment = Alignment(horizontal='left', vertical='center')
        
        # Header row style (gray background, bold)
        header_font = Font(name='Arial', size=11, bold=True)
        header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
               
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15

        # Variable for running point tally
        total_points = 0
        
        # First row: Client code and heading
        ws['A1'] = request.user.get_full_name()  # Client name
        ws['B1'] = request.user.user_number  # Client code (ZČ)
        ws['C1'] = "ČERPÁNÍ ODMĚNY"         # Heading text
        ws['F1'] = "010"                    # Warehouse code
        
        # Apply styles to first row
        for col in ['B', 'C', 'F']:
            ws[f'{col}1'].font = first_row_font
            ws[f'{col}1'].alignment = first_row_alignment
        
        # Row 2 is empty
        
        # Row 3: Headers
        ws['B3'] = "Kód"
        ws['C3'] = "Název"
        ws['D3'] = "Body/ks"
        ws['E3'] = "Cena/ks"
        ws['F3'] = "Množství"
        
        # Apply styles to header row
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws[f'{col}3'].font = header_font
            ws[f'{col}3'].fill = header_fill
            ws[f'{col}3'].alignment = header_alignment
        
        # Get items from the request
        items = request.rewardrequestitem_set.select_related('reward').all()
        
        # Row 4 and onwards: Data rows
        row_num = 4
        for item in items:
            ws[f'B{row_num}'] = item.reward.abra_code
            ws[f'C{row_num}'] = item.reward.name
            ws[f'D{row_num}'] = item.reward.point_cost
            ws[f'E{row_num}'] = 1.0
            ws[f'F{row_num}'] = item.quantity

            # Add point cost to tally
            total_points += item.reward.point_cost * item.quantity
            
            # Apply styles to data cells
            apply_item_styling(row_num)
            
            row_num += 1
        
        # Add a row for BONBOD point item
        ws[f'B{row_num}'] = "BONBOD"
        ws[f'C{row_num}'] = "Bonusový program - čerpání bodů"
        ws[f'D{row_num}'] = ""
        ws[f'E{row_num}'] = 0
        ws[f'F{row_num}'] = total_points

        apply_item_styling(row_num)

        row_num += 1
        
        # Add customer note if it exists
        if request.note:
            row_num += 1  # Add a blank row
            ws[f'B{row_num}'] = "Customer Note:"
            ws[f'B{row_num}'].font = header_font
            
            row_num += 1
            ws.merge_cells(f'B{row_num}:F{row_num}')
            ws[f'B{row_num}'] = request.note
            ws[f'B{row_num}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            # Set row height to accommodate multi-line text
            ws.row_dimensions[row_num].height = 80
            
        # Save the workbook to a BytesIO object
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
        
    except Exception as e:
        # Log error here if needed
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error generating telemarketing export: {str(e)}")
        return None